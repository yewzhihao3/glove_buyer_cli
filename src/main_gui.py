import tkinter as tk
from tkinter import ttk, messagebox
import customtkinter as ctk
import GUI_db
import deepseek_agent
import threading
import csv
import os
from tkinter import filedialog
import db  # Add this import at the top if not present
from concurrent.futures import ThreadPoolExecutor
import queue
import time
import db_apollo
from datetime import datetime, timedelta
from collections import defaultdict, Counter

ctk.set_appearance_mode("light")
ctk.set_default_color_theme("blue")

# Global background task manager
class BackgroundTaskManager:
    def __init__(self):
        self.executor = ThreadPoolExecutor(max_workers=4)
        self.task_queue = queue.Queue()
        self.running = True
        
        # Start background worker thread
        self.worker_thread = threading.Thread(target=self._worker, daemon=True)
        self.worker_thread.start()
    
    def _worker(self):
        """Background worker that processes tasks"""
        while self.running:
            try:
                task = self.task_queue.get(timeout=1)
                if task is None:  # Shutdown signal
                    break
                
                func, args, kwargs, callback, error_callback = task
                try:
                    result = func(*args, **kwargs)
                    if callback:
                        callback(result)
                except Exception as e:
                    if error_callback:
                        error_callback(e)
                    else:
                        print(f"Background task error: {e}")
                finally:
                    self.task_queue.task_done()
            except queue.Empty:
                continue
    
    def submit_task(self, func, callback=None, error_callback=None, *args, **kwargs):
        """Submit a task to run in background"""
        self.task_queue.put((func, args, kwargs, callback, error_callback))
    
    def shutdown(self):
        """Shutdown the task manager"""
        self.running = False
        self.task_queue.put(None)
        self.executor.shutdown(wait=True)

# Global task manager instance
task_manager = BackgroundTaskManager()

# Simple caching system
class Cache:
    def __init__(self):
        self.cache = {}
        self.cache_timestamps = {}
        self.cache_duration = 300  # 5 minutes
    
    def get(self, key):
        if key in self.cache:
            if time.time() - self.cache_timestamps.get(key, 0) < self.cache_duration:
                return self.cache[key]
            else:
                # Cache expired, remove it
                del self.cache[key]
                if key in self.cache_timestamps:
                    del self.cache_timestamps[key]
        return None
    
    def set(self, key, value):
        self.cache[key] = value
        self.cache_timestamps[key] = time.time()
    
    def clear(self):
        self.cache.clear()
        self.cache_timestamps.clear()
    
    def invalidate(self, pattern=None):
        """Invalidate cache entries matching a pattern"""
        if pattern is None:
            self.clear()
        else:
            keys_to_remove = [key for key in self.cache.keys() if pattern in key]
            for key in keys_to_remove:
                del self.cache[key]
                if key in self.cache_timestamps:
                    del self.cache_timestamps[key]

# Global cache instance
cache = Cache()

# Database operation wrappers for background execution
def run_in_background(func, callback=None, error_callback=None, *args, **kwargs):
    """Helper function to run database operations in background"""
    task_manager.submit_task(func, callback, error_callback, *args, **kwargs)

NAV_LABELS = [
    "HS Code",
    "Quick Buyer Search (AI)",
    "AI Buyer Results",
    "Verified Buyer Leads (Apollo)",
    "Buyer List (Apollo)",
    "Export",
    "Settings"
]

class DashboardContent(ctk.CTkFrame):
    def __init__(self, master):
        super().__init__(master, fg_color="#F5F7FA")
        self._build_ui()
        self.load_dashboard_data()
        
    def _build_ui(self):
        # Title
        title = ctk.CTkLabel(self, text="Glove Buyer Intelligence Dashboard", 
                            font=("Poppins", 28, "bold"), text_color="#2E3A59")
        title.pack(pady=(32, 12))
        
        # Simple Stats Frame
        stats_frame = ctk.CTkFrame(self, fg_color="#FFFFFF", corner_radius=16)
        stats_frame.pack(pady=8, padx=32, fill="x")
        stats_frame.grid_columnconfigure((0,1,2,3), weight=1)
        
        # AI Results
        ctk.CTkLabel(stats_frame, text="AI Buyer Results", font=("Poppins", 16), text_color="#6B7C93").grid(row=0, column=0, padx=24, pady=16)
        self.ai_count_label = ctk.CTkLabel(stats_frame, text="Loading...", font=("Poppins", 22, "bold"), text_color="#0078D4")
        self.ai_count_label.grid(row=1, column=0)
        
        # Apollo Contacts
        ctk.CTkLabel(stats_frame, text="Verified Contacts", font=("Poppins", 16), text_color="#6B7C93").grid(row=0, column=1, padx=24, pady=16)
        self.apollo_count_label = ctk.CTkLabel(stats_frame, text="Loading...", font=("Poppins", 22, "bold"), text_color="#4CAF50")
        self.apollo_count_label.grid(row=1, column=1)
        
        # HS Codes
        ctk.CTkLabel(stats_frame, text="HS Codes", font=("Poppins", 16), text_color="#6B7C93").grid(row=0, column=2, padx=24, pady=16)
        self.hs_count_label = ctk.CTkLabel(stats_frame, text="Loading...", font=("Poppins", 22, "bold"), text_color="#FF9800")
        self.hs_count_label.grid(row=1, column=2)
        
        # Companies
        ctk.CTkLabel(stats_frame, text="Companies", font=("Poppins", 16), text_color="#6B7C93").grid(row=0, column=3, padx=24, pady=16)
        self.company_count_label = ctk.CTkLabel(stats_frame, text="Loading...", font=("Poppins", 22, "bold"), text_color="#9C27B0")
        self.company_count_label.grid(row=1, column=3)
        
        # Quick Actions Frame
        actions_frame = ctk.CTkFrame(self, fg_color="#F5F7FA", corner_radius=16)
        actions_frame.pack(pady=24, padx=32, fill="x")
        actions_frame.grid_columnconfigure((0,1,2,3), weight=1)
        
        ctk.CTkButton(actions_frame, text="Buyer Search", fg_color="#0078D4", hover_color="#005A9E", 
                     text_color="#FFFFFF", font=("Poppins", 15, "bold"), corner_radius=8).grid(row=0, column=0, padx=18, pady=18, sticky="ew")
        ctk.CTkButton(actions_frame, text="View Leads", fg_color="#0078D4", hover_color="#005A9E", 
                     text_color="#FFFFFF", font=("Poppins", 15, "bold"), corner_radius=8).grid(row=0, column=1, padx=18, pady=18, sticky="ew")
        ctk.CTkButton(actions_frame, text="HS Code Manager", fg_color="#0078D4", hover_color="#005A9E", 
                     text_color="#FFFFFF", font=("Poppins", 15, "bold"), corner_radius=8).grid(row=0, column=2, padx=18, pady=18, sticky="ew")
        ctk.CTkButton(actions_frame, text="Export", fg_color="#4CAF50", hover_color="#388E3C", 
                     text_color="#FFFFFF", font=("Poppins", 15, "bold"), corner_radius=8).grid(row=0, column=3, padx=18, pady=18, sticky="ew")
        
        # Charts and Activity Frame
        charts_activity_frame = ctk.CTkFrame(self, fg_color="transparent")
        charts_activity_frame.pack(pady=8, padx=32, fill="both", expand=True)
        charts_activity_frame.grid_columnconfigure((0, 1), weight=1)
        
        # Country Distribution Chart Frame
        chart_frame = ctk.CTkFrame(charts_activity_frame, fg_color="#FFFFFF", corner_radius=16)
        chart_frame.grid(row=0, column=0, padx=(0, 8), pady=0, sticky="ew")
        ctk.CTkLabel(chart_frame, text="🌍 Company Distribution by Country", font=("Poppins", 18, "bold"), text_color="#2E3A59").pack(pady=(16,8))
        
        # Chart text area
        self.chart_text = ctk.CTkTextbox(chart_frame, font=("Consolas", 11), 
                                        fg_color="#F8F9FA", text_color="#2E3A59", height=200)
        self.chart_text.pack(fill="both", expand=True, padx=16, pady=(0, 16))
        
        # Recent Activity Frame
        recent_frame = ctk.CTkFrame(charts_activity_frame, fg_color="#FFFFFF", corner_radius=16)
        recent_frame.grid(row=0, column=1, padx=(8, 0), pady=0, sticky="ew")
        ctk.CTkLabel(recent_frame, text="Recent Activity", font=("Poppins", 18, "bold"), text_color="#2E3A59").pack(pady=(16,8))
        
        # Activity text area
        self.activity_text = ctk.CTkTextbox(recent_frame, font=("Consolas", 12), 
                                           fg_color="#F8F9FA", text_color="#2E3A59", height=200)
        self.activity_text.pack(fill="both", expand=True, padx=16, pady=(0, 16))
        

        

        

        
    def load_dashboard_data(self):
        """Load dashboard data directly without background tasks"""
        try:
            print("Starting dashboard data load...")
            import GUI_db
            
            # Initialize database tables first
            GUI_db.init_db()
            GUI_db.init_apollo_db()
            GUI_db.init_deepseek_results_table()
            
            print("Database tables initialized")
            
            # Get all data with error handling for each
            try:
                ai_results = GUI_db.get_all_deepseek_results()
                print(f"AI results loaded: {len(ai_results)}")
            except Exception as e:
                print(f"Error loading AI results: {e}")
                ai_results = []
            
            try:
                apollo_contacts = GUI_db.get_all_contacts()
                print(f"Apollo contacts loaded: {len(apollo_contacts)}")
            except Exception as e:
                print(f"Error loading Apollo contacts: {e}")
                apollo_contacts = []
            
            try:
                hs_codes = GUI_db.get_all_hs_codes()
                print(f"HS codes loaded: {len(hs_codes)}")
            except Exception as e:
                print(f"Error loading HS codes: {e}")
                hs_codes = []
            
            try:
                companies = GUI_db.get_all_companies()
                print(f"Companies loaded: {len(companies)}")
            except Exception as e:
                print(f"Error loading companies: {e}")
                companies = []
            
            print("All data loaded successfully")
            
            data = {
                'counts': {
                    'ai': len(ai_results),
                    'apollo': len(apollo_contacts),
                    'hs': len(hs_codes),
                    'companies': len(companies)
                },
                'activity': self._generate_simple_activity(ai_results, apollo_contacts, hs_codes),
                'chart': self._generate_country_chart(ai_results, apollo_contacts)
            }
            
            self._update_dashboard_ui(data)
            
        except Exception as e:
            print(f"Error loading dashboard data: {e}")
            import traceback
            traceback.print_exc()
            # Show error state
            self._update_dashboard_ui({
                'counts': {'ai': 0, 'apollo': 0, 'hs': 0, 'companies': 0},
                'activity': f"Error loading dashboard data: {e}",
                'chart': "Error loading chart data"
            })
        
    def _update_dashboard_ui(self, data):
        """Update dashboard UI with loaded data"""
        # Update statistics
        self.ai_count_label.configure(text=str(data['counts']['ai']))
        self.apollo_count_label.configure(text=str(data['counts']['apollo']))
        self.hs_count_label.configure(text=str(data['counts']['hs']))
        self.company_count_label.configure(text=str(data['counts']['companies']))
        
        # Update chart
        self.chart_text.delete("1.0", "end")
        self.chart_text.insert("1.0", data['chart'])
        
        # Update activity
        self.activity_text.delete("1.0", "end")
        self.activity_text.insert("1.0", data['activity'])
    
    def _generate_simple_activity(self, ai_results, contacts, hs_codes):
        """Generate simple activity summary"""
        activity = "Recent Activity Summary:\n\n"
        
        # Show recent AI results
        if ai_results:
            recent_ai = ai_results[-5:]  # Last 5
            activity += f"🤖 Recent AI Searches ({len(ai_results)} total):\n"
            for result in recent_ai:
                activity += f"  • {result.get('keyword', '')} in {result.get('country', '')}\n"
            activity += "\n"
        
        # Show recent contacts
        if contacts:
            recent_contacts = contacts[-3:]  # Last 3
            activity += f"👥 Recent Contacts ({len(contacts)} total):\n"
            for contact in recent_contacts:
                activity += f"  • {contact.get('name', '')} - {contact.get('company_name', '')}\n"
            activity += "\n"
        
        # Show recent HS codes
        if hs_codes:
            recent_hs = hs_codes[-3:]  # Last 3
            activity += f"🏷️ Recent HS Codes ({len(hs_codes)} total):\n"
            for hs_code in recent_hs:
                activity += f"  • {hs_code.get('hs_code', '')} - {hs_code.get('description', '')}\n"
        
        if not ai_results and not contacts and not hs_codes:
            activity += "No recent activity found."
        
        return activity
    
    def _generate_country_chart(self, ai_results, contacts):
        """Generate country distribution chart"""
        from collections import Counter
        import GUI_db
        
        # Collect all countries from AI results and contacts
        all_countries = []
        
        # From AI results
        for result in ai_results:
            country = result.get('company_country', 'Unknown')
            if country and country != 'Unknown':
                all_countries.append(country)
        
        # From contacts - get country from companies table
        try:
            companies = GUI_db.get_all_companies()
            company_countries = {company['id']: company.get('country', 'Unknown') for company in companies}
            
            for contact in contacts:
                company_id = contact.get('company_id')
                if company_id and company_id in company_countries:
                    country = company_countries[company_id]
                    if country and country != 'Unknown':
                        all_countries.append(country)
        except Exception as e:
            print(f"Error getting company countries: {e}")
        
        if not all_countries:
            return "No country data available"
        
        # Count countries
        country_counts = Counter(all_countries)
        
        # Generate chart
        chart = "🌍 Company Distribution by Country\n"
        chart += "=" * 40 + "\n\n"
        
        total = len(all_countries)
        max_count = max(country_counts.values()) if country_counts else 1
        
        # Sort by count (descending)
        sorted_countries = sorted(country_counts.items(), key=lambda x: x[1], reverse=True)
        
        for i, (country, count) in enumerate(sorted_countries[:10], 1):  # Show top 10
            percentage = round((count / total) * 100, 1)
            bar_length = int((count / max_count) * 30)  # Scale bar to max 30 characters
            bar = "█" * bar_length + "░" * (30 - bar_length)
            
            chart += f"{i:2d}. {country:<20} {count:3d} ({percentage:4.1f}%)\n"
            chart += f"    {bar}\n\n"
        
        chart += f"\n📊 Total Companies: {total}"
        chart += f"\n🏆 Top Country: {sorted_countries[0][0] if sorted_countries else 'N/A'}"
        
        return chart
        

    

    
    def _format_time_ago(self, date_str):
        """Format date as time ago"""
        if not date_str:
            return "Unknown"
        
        try:
            date = self._parse_date(date_str)
            now = datetime.now()
            diff = now - date
            
            if diff.days > 0:
                return f"{diff.days} day{'s' if diff.days != 1 else ''} ago"
            elif diff.seconds > 3600:
                hours = diff.seconds // 3600
                return f"{hours} hour{'s' if hours != 1 else ''} ago"
            else:
                minutes = diff.seconds // 60
                return f"{minutes} minute{'s' if minutes != 1 else ''} ago"
        except:
            return "Unknown"
    
    def _parse_date(self, date_str):
        """Parse date string to datetime object"""
        if not date_str:
            return datetime.min
        try:
            for fmt in ['%Y-%m-%dT%H:%M:%S.%f', '%Y-%m-%dT%H:%M:%S', '%Y-%m-%d %H:%M:%S', '%Y-%m-%d']:
                try:
                    return datetime.strptime(date_str, fmt)
                except ValueError:
                    continue
            return datetime.min
        except:
            return datetime.min
    

    


class BuyerSearchPage(ctk.CTkFrame):
    def __init__(self, master):
        super().__init__(master, fg_color="#F5F7FA")
        self._build_ui()
        self.search_results = []
        self.worker = None
        # Load countries synchronously after widgets are created
        self.load_countries()

    def _build_ui(self):
        # Title
        title_frame = ctk.CTkFrame(self, fg_color="#F5F7FA")
        title_frame.pack(fill="x", pady=(24, 16), padx=32)
        ctk.CTkLabel(title_frame, text="Quick Buyer Search (AI)", font=("Poppins", 24, "bold"), text_color="#2E3A59").pack()
        
        # Search Parameters Card
        search_card = ctk.CTkFrame(self, fg_color="#FFFFFF", corner_radius=16)
        search_card.pack(fill="x", padx=32, pady=(0, 16))
        
        search_content = ctk.CTkFrame(search_card, fg_color="transparent")
        search_content.pack(fill="x", padx=20, pady=18)
        
        ctk.CTkLabel(search_content, text="Search Parameters", font=("Poppins", 18, "bold"), text_color="#0078D4").pack(anchor="w", pady=(0, 12))
        
        # Parameters grid
        param_frame = ctk.CTkFrame(search_content, fg_color="transparent")
        param_frame.pack(fill="x", pady=8)
        param_frame.grid_columnconfigure((0, 1), weight=1)
        
        # Country selection
        country_frame = ctk.CTkFrame(param_frame, fg_color="transparent")
        country_frame.grid(row=0, column=0, sticky="ew", padx=(0, 8), pady=4)
        ctk.CTkLabel(country_frame, text="Country:", font=("Poppins", 15), text_color="#6B7C93").pack(anchor="w")
        self.country_var = tk.StringVar(value="Select Country")
        country_display_frame = ctk.CTkFrame(country_frame, fg_color="transparent")
        country_display_frame.pack(fill="x", pady=(4, 0))
        self.country_display = ctk.CTkEntry(country_display_frame, textvariable=self.country_var, state="readonly", width=200, font=("Poppins", 15))
        self.country_display.pack(side="left")
        select_country_btn = ctk.CTkButton(country_display_frame, text="Select", fg_color="#4CAF50", hover_color="#388E3C", text_color="#FFFFFF", font=("Poppins", 12, "bold"), width=60, height=32, command=self.open_country_selector)
        select_country_btn.pack(side="left", padx=(8, 0))
        
        # HS Code selection
        hs_frame = ctk.CTkFrame(param_frame, fg_color="transparent")
        hs_frame.grid(row=0, column=1, sticky="ew", padx=(8, 0), pady=4)
        ctk.CTkLabel(hs_frame, text="HS Code:", font=("Poppins", 15), text_color="#6B7C93").pack(anchor="w")
        self.hs_var = tk.StringVar(value="Select HS Code")
        hs_display_frame = ctk.CTkFrame(hs_frame, fg_color="transparent")
        hs_display_frame.pack(fill="x", pady=(4, 0))
        self.hs_display = ctk.CTkEntry(hs_display_frame, textvariable=self.hs_var, state="readonly", width=200, font=("Poppins", 15))
        self.hs_display.pack(side="left")
        select_hs_btn = ctk.CTkButton(hs_display_frame, text="Select", fg_color="#4CAF50", hover_color="#388E3C", text_color="#FFFFFF", font=("Poppins", 12, "bold"), width=60, height=32, command=self.open_hs_code_selector)
        select_hs_btn.pack(side="left", padx=(8, 0))
        
        # Keyword selection
        keyword_frame = ctk.CTkFrame(param_frame, fg_color="transparent")
        keyword_frame.grid(row=1, column=0, sticky="ew", padx=(0, 8), pady=4)
        ctk.CTkLabel(keyword_frame, text="Product Keyword:", font=("Poppins", 15), text_color="#6B7C93").pack(anchor="w")
        self.keyword_var = tk.StringVar(value="Select Keyword")
        keyword_display_frame = ctk.CTkFrame(keyword_frame, fg_color="transparent")
        keyword_display_frame.pack(fill="x", pady=(4, 0))
        self.keyword_display = ctk.CTkEntry(keyword_display_frame, textvariable=self.keyword_var, state="readonly", width=200, font=("Poppins", 15))
        self.keyword_display.pack(side="left")
        select_keyword_btn = ctk.CTkButton(keyword_display_frame, text="Select", fg_color="#4CAF50", hover_color="#388E3C", text_color="#FFFFFF", font=("Poppins", 12, "bold"), width=60, height=32, command=self.open_keyword_selector)
        select_keyword_btn.pack(side="left", padx=(8, 0))
        
        # Custom keyword entry
        custom_frame = ctk.CTkFrame(param_frame, fg_color="transparent")
        custom_frame.grid(row=1, column=1, sticky="ew", padx=(8, 0), pady=4)
        ctk.CTkLabel(custom_frame, text="Custom Keyword:", font=("Poppins", 15), text_color="#6B7C93").pack(anchor="w")
        self.custom_keyword_var = tk.StringVar()
        self.custom_keyword_entry = ctk.CTkEntry(custom_frame, textvariable=self.custom_keyword_var, placeholder_text="Enter custom keyword...", width=200, font=("Poppins", 15), state="disabled")
        self.custom_keyword_entry.pack(fill="x", pady=(4, 0))
        
        # Search button
        button_frame = ctk.CTkFrame(param_frame, fg_color="transparent")
        button_frame.grid(row=2, column=0, columnspan=2, sticky="ew", pady=(12, 0))
        
        self.search_btn = ctk.CTkButton(button_frame, text="Search Buyers with AI", fg_color="#0078D4", hover_color="#005A9E", text_color="#FFFFFF", font=("Poppins", 15, "bold"), corner_radius=8, command=self.perform_search)
        self.search_btn.pack(side="right")
        
        # Results Card
        results_card = ctk.CTkFrame(self, fg_color="#FFFFFF", corner_radius=16)
        results_card.pack(fill="both", expand=True, padx=32, pady=(0, 24))
        
        results_content = ctk.CTkFrame(results_card, fg_color="transparent")
        results_content.pack(fill="both", expand=True, padx=20, pady=18)
        
        # Info and action buttons row
        info_action_frame = ctk.CTkFrame(results_content, fg_color="transparent")
        info_action_frame.pack(fill="x", pady=(0, 12))
        
        self.results_info_label = ctk.CTkLabel(info_action_frame, text="", font=("Poppins", 15), text_color="#0078D4", fg_color="#EAF3FF", corner_radius=8)
        self.results_info_label.pack(side="left", padx=(0, 16))
        
        # Action buttons
        export_btn = ctk.CTkButton(info_action_frame, text="Export Results", fg_color="#0078D4", hover_color="#005A9E", text_color="#FFFFFF", font=("Poppins", 15, "bold"), corner_radius=8, width=180, height=40, command=self.export_results)
        export_btn.pack(side="right", padx=(0, 8))
        
        # Table
        table_frame = ctk.CTkFrame(results_content, fg_color="#FFFFFF", corner_radius=12)
        table_frame.pack(fill="both", expand=True, pady=(0, 8))
        
        style = ttk.Style()
        style.configure("BuyerSearch.Treeview", font=("Poppins", 13), rowheight=32, background="#FFFFFF", fieldbackground="#FFFFFF", foreground="#2E3A59")
        style.configure("BuyerSearch.Treeview.Heading", font=("Poppins", 14, "bold"), background="#F5F7FA", foreground="#2E3A59")
        style.map("BuyerSearch.Treeview", background=[("selected", "#1E3A8A")], foreground=[("selected", "#FFFFFF")])
        
        # Add horizontal scrollbar
        xscroll = tk.Scrollbar(table_frame, orient="horizontal")
        xscroll.pack(side="bottom", fill="x")
        
        self.table = ttk.Treeview(
            table_frame,
            columns=("no", "company", "country", "search_country", "website", "description"),
            show="headings",
            style="BuyerSearch.Treeview",
            xscrollcommand=xscroll.set
        )
        self.table.heading("no", text="No.")
        self.table.heading("company", text="Company Name")
        self.table.heading("country", text="Country")
        self.table.heading("search_country", text="Search Country")
        self.table.heading("website", text="Website")
        self.table.heading("description", text="Description")
        self.table.column("no", width=60, anchor="center")
        self.table.column("company", width=200, anchor="w")
        self.table.column("country", width=120, anchor="center")
        self.table.column("search_country", width=140, anchor="center")
        self.table.column("website", width=200, anchor="w")
        self.table.column("description", width=300, anchor="w")
        self.table.pack(fill="both", expand=True, padx=8, pady=8)
        xscroll.config(command=self.table.xview)
        
        # Progress bar
        self.progress_bar = ctk.CTkProgressBar(results_content, orientation="horizontal", mode="indeterminate", width=320)
        self.progress_bar.pack(pady=8)
        self.progress_bar.set(0)
        self.progress_bar.pack_forget()  # Hide initially
        
        # Initialize data
        self.load_keywords()
        self.load_hs_codes()
        
        # Load countries after a short delay to ensure UI is ready
        self.after(100, self.load_countries)

    def load_countries(self):
        """Load available countries from database that have HS codes"""
        # Check cache first
        cached_countries = cache.get("buyer_search_countries")
        if cached_countries:
            self.country_list = cached_countries
            return
        try:
            hs_codes = GUI_db.get_all_hs_codes()
            countries_with_hs_codes = set()
            for code in hs_codes:
                if code.get('country'):
                    countries_with_hs_codes.add(code['country'])
            countries = ["Select Country"] + sorted(list(countries_with_hs_codes))
            self.country_list = countries
            cache.set("buyer_search_countries", countries)
        except Exception as e:
            print(f"Error loading countries: {e}")
            self.country_list = ["Select Country"]

    def _update_countries_ui(self, countries):
        """Update countries list on main thread"""
        self.country_list = countries
        # Cache the result
        cache.set("buyer_search_countries", countries)
        print(f"[QUICK DEBUG] Updated countries UI: {len(countries)} countries")
    
    def _set_default_countries(self):
        """Set default countries on main thread"""
        self.country_list = ["Select Country"]

    def load_keywords(self):
        """Load keyword options from file and add custom option"""
        try:
            keyword_file = os.path.join(os.path.dirname(__file__), '..', 'prompts', 'keyword_options.txt')
            if os.path.exists(keyword_file):
                with open(keyword_file, 'r', encoding='utf-8') as f:
                    keywords = [line.strip() for line in f if line.strip()]
                self.keyword_list = ["Select Keyword"] + keywords + ["Custom Keyword"]
            else:
                self.keyword_list = ["Select Keyword", "Custom Keyword"]
        except Exception as e:
            print(f"Error loading keywords: {e}")
            self.keyword_list = ["Select Keyword", "Custom Keyword"]

    def load_hs_codes(self):
        """Load HS codes based on selected country"""
        try:
            # Default to empty list, will be populated when country is selected
            self.hs_code_list = ["Select HS Code"]
        except Exception as e:
            print(f"Error loading HS codes: {e}")
            self.hs_code_list = ["Select HS Code"]

    def open_country_selector(self):
        """Open country selection dialog"""
        dialog = CountrySelectorDialog(self, self.country_list)
        self.wait_window(dialog)
        selected_country = dialog.get_selected()
        if selected_country and selected_country != "Select Country":
            self.country_var.set(selected_country)
            # Update HS codes for selected country
            self.update_hs_codes_for_country(selected_country)
            # Reset keyword selection when country changes
            self.reset_keyword_selection()

    def open_hs_code_selector(self):
        """Open HS code selection dialog"""
        country = self.country_var.get().strip()
        if country == "Select Country" or not country:
            messagebox.showwarning("Select Country First", "Please select a country first to see available HS codes.")
            return
        
        # Show loading indicator
        loading_dialog = ctk.CTkToplevel(self)
        loading_dialog.title("Loading")
        loading_dialog.geometry("300x100")
        loading_dialog.grab_set()
        loading_dialog.transient(self.winfo_toplevel())
        
        ctk.CTkLabel(loading_dialog, text="Loading HS codes...", font=("Poppins", 14)).pack(pady=20)
        progress_bar = ctk.CTkProgressBar(loading_dialog, mode="indeterminate")
        progress_bar.pack(pady=10)
        progress_bar.start()
        
        def load_hs_codes_task():
            try:
                hs_codes = GUI_db.get_hs_codes_by_country(country)
                hs_options = ["Select HS Code"]
                for code in hs_codes:
                    hs_options.append(f"{code['hs_code']} - {code['description']}")
                return hs_options
            except Exception as e:
                print(f"Error loading HS codes for country: {e}")
                return None
        
        def on_hs_codes_loaded(hs_options):
            # Schedule UI update on main thread
            self.after(0, lambda: self._handle_hs_codes_loaded(hs_options, loading_dialog, country))
        
        def on_error(error):
            # Schedule UI update on main thread
            self.after(0, lambda: self._handle_hs_codes_error(error, loading_dialog, country))
        
        # Run in background
        run_in_background(load_hs_codes_task, on_hs_codes_loaded, on_error)
    
    def _handle_hs_codes_loaded(self, hs_options, loading_dialog, country):
        """Handle HS codes loaded on main thread"""
        loading_dialog.destroy()
        if hs_options is None:
            messagebox.showerror("Error", f"Error loading HS codes for {country}")
            return
        
        dialog = HSCodeSelectorDialog(self, hs_options)
        self.wait_window(dialog)
        selected_hs = dialog.get_selected()
        if selected_hs and selected_hs != "Select HS Code":
            self.hs_var.set(selected_hs)
    
    def _handle_hs_codes_error(self, error, loading_dialog, country):
        """Handle HS codes error on main thread"""
        loading_dialog.destroy()
        print(f"Error loading HS codes for country: {error}")
        messagebox.showerror("Error", f"Error loading HS codes for {country}")

    def open_keyword_selector(self):
        """Open keyword selection dialog"""
        dialog = KeywordSelectorDialog(self, self.keyword_list)
        self.wait_window(dialog)
        selected_keyword = dialog.get_selected()
        if selected_keyword and selected_keyword != "Select Keyword":
            if selected_keyword == "Custom Keyword":
                # Enable custom keyword entry and clear other selections
                self.keyword_var.set("Custom Keyword")
                self.custom_keyword_entry.configure(state="normal")
                self.custom_keyword_entry.focus()
            else:
                # Disable custom keyword entry and set the selected keyword
                self.keyword_var.set(selected_keyword)
                self.custom_keyword_entry.configure(state="disabled")
                self.custom_keyword_var.set("")

    def reset_keyword_selection(self):
        """Reset keyword selection to default state"""
        self.keyword_var.set("Select Keyword")
        self.custom_keyword_entry.configure(state="disabled")
        self.custom_keyword_var.set("")

    def update_hs_codes_for_country(self, country):
        """Update HS codes when country changes"""
        try:
            if country == "Select Country":
                self.hs_code_list = ["Select HS Code"]
            else:
                hs_codes = GUI_db.get_hs_codes_by_country(country)
                self.hs_code_list = ["Select HS Code"] + [f"{code['hs_code']} - {code['description']}" for code in hs_codes]
            self.hs_var.set("Select HS Code")
        except Exception as e:
            print(f"Error updating HS codes for country: {e}")

    def refresh_buyer_search_data(self):
        """Refresh BuyerSearchPage data after new HS codes are added"""
        # Invalidate cache
        cache.invalidate("buyer_search_countries")
        cache.invalidate("hs_codes")
        self.load_countries()
    
    def refresh_ai_buyer_results_page(self):
        """Refresh AI Buyer Results page after new DeepSeek results are added"""
        try:
            # Find the main app instance to access other pages
            main_app = self.winfo_toplevel()
            if hasattr(main_app, 'pages'):
                # Refresh AI Buyer Results page (index 2)
                ai_buyer_results_page = main_app.pages[2]  # Index 2 is DeepSeekBuyerResultsPage
                if hasattr(ai_buyer_results_page, 'refresh_country_list'):
                    ai_buyer_results_page.refresh_country_list()
        except Exception as e:
            print(f"Error refreshing AI Buyer Results page: {e}")

    def perform_search(self):
        """Perform AI buyer search"""
        country = self.country_var.get().strip()
        hs_selection = self.hs_var.get().strip()
        keyword = self.keyword_var.get().strip()
        custom_keyword = self.custom_keyword_var.get().strip()
        
        # Validation
        if country == "Select Country" or not country:
            messagebox.showwarning("Missing Country", "Please select a country.")
            return
            
        if hs_selection == "Select HS Code" or not hs_selection:
            messagebox.showwarning("Missing HS Code", "Please select an HS code.")
            return
            
        # Check if custom keyword is selected but no custom keyword entered
        if keyword == "Custom Keyword":
            if not custom_keyword:
                messagebox.showwarning("Missing Custom Keyword", "Please enter a custom keyword.")
                return
            final_keyword = custom_keyword
        elif keyword == "Select Keyword" or not keyword:
            messagebox.showwarning("Missing Keyword", "Please select a keyword or choose custom keyword.")
            return
        else:
            final_keyword = keyword
        
        # Extract HS code from selection
        hs_code = hs_selection.split(" - ")[0] if " - " in hs_selection else hs_selection
        
        # Start search
        self.search_btn.configure(state="disabled")
        self.progress_bar.pack(pady=8)
        self.progress_bar.start()
        
        def run_ai_search():
            try:
                import deepseek_agent
                import db
                
                print(f"[DEBUG] Starting DeepSeek search with parameters:")
                print(f"[DEBUG] HS Code: {hs_code}")
                print(f"[DEBUG] Keyword: {final_keyword}")
                print(f"[DEBUG] Country: {country}")
                
                # Query DeepSeek for buyers
                print("[DEBUG] Calling deepseek_agent.query_deepseek...")
                result = deepseek_agent.query_deepseek(hs_code, final_keyword, country, [])
                print(f"[DEBUG] DeepSeek raw result received: {len(str(result))} characters")
                print(f"[DEBUG] First 500 chars of result: {str(result)[:500]}...")
                
                # Parse results
                print("[DEBUG] Parsing DeepSeek output...")
                companies = db.parse_deepseek_output(result)
                print(f"[DEBUG] Parsed {len(companies) if companies else 0} companies from result")
                
                # Debug: Show what fields each company has
                if companies:
                    print("[DEBUG] Sample company data:")
                    for i, company in enumerate(companies[:3]):  # Show first 3 companies
                        print(f"[DEBUG] Company {i+1}:")
                        print(f"[DEBUG]   Name: {company.get('company_name', 'MISSING')}")
                        print(f"[DEBUG]   Country: {company.get('company_country', 'MISSING')}")
                        print(f"[DEBUG]   Website: {company.get('company_website_link', 'MISSING')}")
                        print(f"[DEBUG]   Description: {company.get('description', 'MISSING')[:100]}...")
                
                # Save to database
                if companies:
                    print("[DEBUG] Saving companies to database...")
                    # Use the new deepseek_buyer_search_results table
                    saved_count = GUI_db.insert_deepseek_results(hs_code, final_keyword, country, companies)
                    print(f"[DEBUG] Saved {saved_count} companies to deepseek_buyer_search_results table")
                
                def on_complete():
                    self.search_btn.configure(state="normal")
                    self.progress_bar.stop()
                    self.progress_bar.pack_forget()
                    
                    if not companies:
                        self.results_info_label.configure(text="")
                        messagebox.showinfo("No Results", "No buyers found for the specified criteria.")
                        return
                    
                    self.search_results = companies
                    self.populate_table(companies)
                    self.results_info_label.configure(text=f"Found {len(companies)} potential buyers")
                    
                    # Auto-refresh country list after successful search
                    self.load_countries()
                    
                    # Refresh AI Buyer Results page if it exists
                    self.refresh_ai_buyer_results_page()
                    
                    messagebox.showinfo("Search Complete", f"Found and saved {len(companies)} potential buyers to the database.")
                
                self.after(0, on_complete)
                
            except Exception as search_error:
                print(f"[DEBUG] Error during AI search: {search_error}")
                print(f"[DEBUG] Error type: {type(search_error)}")
                import traceback
                print(f"[DEBUG] Full traceback: {traceback.format_exc()}")
                
                def on_error(error=search_error):
                    self.search_btn.configure(state="normal")
                    self.progress_bar.stop()
                    self.progress_bar.pack_forget()
                    messagebox.showerror("Search Error", f"Error during AI search: {str(error)}")
                self.after(0, on_error)
        
        threading.Thread(target=run_ai_search, daemon=True).start()

    def populate_table(self, data):
        """Populate the results table"""
        for row in self.table.get_children():
            self.table.delete(row)
        search_country = self.country_var.get()
        for i, company in enumerate(data):
            self.table.insert("", "end", values=(
                str(i + 1),
                company.get('company_name', ''),
                company.get('company_country', ''),
                search_country,
                company.get('company_website_link', ''),
                company.get('description', '')
            ))

    def export_results(self):
        """Export results to CSV"""
        if not self.search_results:
            messagebox.showwarning("No Data", "No results to export.")
            return
            
        try:
            # Generate default filename
            country = self.country_var.get()
            hs_code = self.hs_var.get().split(" - ")[0] if " - " in self.hs_var.get() else self.hs_var.get()
            default_filename = f"{country}_{hs_code}_buyers.csv"
            
            filename = filedialog.asksaveasfilename(
                defaultextension=".csv",
                filetypes=[("CSV files", "*.csv"), ("All files", "*.*")],
                title="Export Buyer Search Results",
                initialfile=default_filename
            )
            
            if filename:
                with open(filename, 'w', newline='', encoding='utf-8') as csvfile:
                    writer = csv.writer(csvfile)
                    # Write headers
                    headers = ["No.", "Company Name", "Country", "Website", "Description"]
                    writer.writerow(headers)
                    
                    # Write data
                    for i, company in enumerate(self.search_results):
                        writer.writerow([
                            str(i + 1),
                            company.get('company_name', ''),
                            company.get('company_country', ''),
                            company.get('company_website_link', ''),
                            company.get('description', '')
                        ])
                
                messagebox.showinfo("Export Complete", f"Results exported to {filename}")
                
        except Exception as e:
            messagebox.showerror("Export Error", f"Error exporting results: {e}")


class DeepSeekBuyerResultsPage(ctk.CTkFrame):
    def __init__(self, master):
        super().__init__(master, fg_color="#F5F7FA")
        self._build_ui()
        self.load_countries()
        self.populate_table()

    def _build_ui(self):
        # Title
        title_frame = ctk.CTkFrame(self, fg_color="#F5F7FA")
        title_frame.pack(fill="x", pady=(24, 16), padx=32)
        ctk.CTkLabel(title_frame, text="AI Buyer Results (DeepSeek)", font=("Poppins", 24, "bold"), text_color="#2E3A59").pack(side="left")
        
        # Search/filter row
        filter_frame = ctk.CTkFrame(self, fg_color="#F5F7FA")
        filter_frame.pack(fill="x", pady=(18, 0), padx=32)
        
        ctk.CTkLabel(filter_frame, text="Search:", font=("Poppins", 15), text_color="#6B7C93").pack(side="left", padx=(0, 8))
        self.search_var = tk.StringVar()
        self.search_var.trace("w", self.on_search_change)
        search_entry = ctk.CTkEntry(filter_frame, textvariable=self.search_var, placeholder_text="Search companies, descriptions, HS codes...", width=300, font=("Poppins", 15))
        search_entry.pack(side="left", padx=(0, 12))
        
        ctk.CTkLabel(filter_frame, text="Country:", font=("Poppins", 15), text_color="#6B7C93").pack(side="left", padx=(16, 8))
        self.country_var = tk.StringVar(value="All")
        country_frame = ctk.CTkFrame(filter_frame, fg_color="transparent")
        country_frame.pack(side="left", padx=(0, 12))
        
        # Country display (read-only)
        self.country_display = ctk.CTkEntry(country_frame, textvariable=self.country_var, state="readonly", width=160, font=("Poppins", 15))
        self.country_display.pack(side="left")
        
        # Select country button
        select_country_btn = ctk.CTkButton(country_frame, text="Select", fg_color="#4CAF50", hover_color="#388E3C", text_color="#FFFFFF", font=("Poppins", 12, "bold"), width=60, height=32, command=self.open_country_selector)
        select_country_btn.pack(side="left", padx=(8, 0))
        
        refresh_btn = ctk.CTkButton(filter_frame, text="Refresh", fg_color="#4CAF50", hover_color="#388E3C", text_color="#FFFFFF", font=("Poppins", 15), corner_radius=8, width=100, command=self.populate_table)
        refresh_btn.pack(side="left", padx=(0, 8))
        
        clear_btn = ctk.CTkButton(filter_frame, text="Clear Search", fg_color="#B0BEC5", hover_color="#90A4AE", text_color="#FFFFFF", font=("Poppins", 15), corner_radius=8, width=120, command=self.clear_search)
        clear_btn.pack(side="left")

        # Table frame
        table_frame = ctk.CTkFrame(self, fg_color="#FFFFFF", corner_radius=16)
        table_frame.pack(fill="both", expand=True, padx=32, pady=(18, 0))
        
        style = ttk.Style()
        style.configure("DeepSeek.Treeview", font=("Poppins", 13), rowheight=32, background="#FFFFFF", fieldbackground="#FFFFFF", foreground="#2E3A59")
        style.configure("DeepSeek.Treeview.Heading", font=("Poppins", 14, "bold"), background="#F5F7FA", foreground="#0078D4")
        style.map("DeepSeek.Treeview", background=[("selected", "#1E3A8A")], foreground=[("selected", "#FFFFFF")])
        
        self.table = ttk.Treeview(table_frame, columns=("id", "hs_code", "keyword", "search_country", "company_name", "company_country", "website", "description"), show="headings", style="DeepSeek.Treeview")
        self.table.heading("id", text="ID")
        self.table.heading("hs_code", text="HS Code")
        self.table.heading("keyword", text="Keyword")
        self.table.heading("search_country", text="Search Country")
        self.table.heading("company_name", text="Company Name")
        self.table.heading("company_country", text="Company Country")
        self.table.heading("website", text="Website")
        self.table.heading("description", text="Description")
        
        self.table.column("id", width=60, anchor="center")
        self.table.column("hs_code", width=100, anchor="center")
        self.table.column("keyword", width=120, anchor="w")
        self.table.column("search_country", width=120, anchor="center")
        self.table.column("company_name", width=250, anchor="w")
        self.table.column("company_country", width=120, anchor="center")
        self.table.column("website", width=200, anchor="w")
        self.table.column("description", width=300, anchor="w")
        
        self.table.pack(fill="both", expand=True, padx=8, pady=8)
        self._add_table_sorting()

        # Action buttons below table
        action_frame = ctk.CTkFrame(self, fg_color="#F5F7FA")
        action_frame.pack(fill="x", padx=32, pady=(8, 24))
        
        export_btn = ctk.CTkButton(action_frame, text="Export Results", fg_color="#0078D4", hover_color="#005A9E", text_color="#FFFFFF", font=("Poppins", 15, "bold"), corner_radius=8, width=150, command=self.export_results)
        export_btn.pack(side="right", padx=(0, 12))
        
        edit_btn = ctk.CTkButton(action_frame, text="Edit Selected", fg_color="#4CAF50", hover_color="#388E3C", text_color="#FFFFFF", font=("Poppins", 15), corner_radius=8, width=140, command=self.edit_selected)
        edit_btn.pack(side="right", padx=(0, 12))
        
        delete_btn = ctk.CTkButton(action_frame, text="Delete Selected", fg_color="#F44336", hover_color="#D32F2F", text_color="#FFFFFF", font=("Poppins", 15), corner_radius=8, width=140, command=self.delete_selected)
        delete_btn.pack(side="right", padx=(0, 12))

    def populate_table(self):
        """Populate the table with DeepSeek results, sorted by ID ascending by default"""
        try:
            search_term = self.search_var.get().strip()
            selected_country = self.country_var.get().strip()
            # Get all results first
            if search_term:
                results = GUI_db.get_deepseek_results_by_search(search_term)
            else:
                results = GUI_db.get_all_deepseek_results()
            # Filter by country if not "All"
            if selected_country and selected_country != "All":
                filtered_results = []
                for result in results:
                    if result.get('company_country', '').strip() == selected_country:
                        filtered_results.append(result)
                results = filtered_results
            # Sort by id ascending
            results = sorted(results, key=lambda x: x.get('id', 0))
            
            # Update UI directly on main thread
            self._update_table_ui(results)
            # Also refresh country list to include any new countries
            self.refresh_country_list()
            
        except Exception as error:
            print(f"Error loading DeepSeek results: {error}")
            self._clear_table_ui()

    def _add_table_sorting(self):
        """Enable sorting by clicking column headers"""
        for col in self.table['columns']:
            self.table.heading(col, command=lambda c=col: self._sort_by_column(c, False))

    def _sort_by_column(self, col, descending):
        data = [(self.table.set(child, col), child) for child in self.table.get_children('')]
        # Try to convert to int for id column
        if col == 'id':
            def sort_key(item):
                try:
                    return int(item[0]) if str(item[0]).isdigit() else 0
                except (ValueError, AttributeError):
                    return 0
            data.sort(key=sort_key, reverse=descending)
        else:
            data.sort(reverse=descending)
        for index, (val, child) in enumerate(data):
            self.table.move(child, '', index)
        # Reverse sort next time
        self.table.heading(col, command=lambda: self._sort_by_column(col, not descending))

    def refresh_country_list(self):
        """Refresh the country list from database"""
        self.load_countries()
    
    def _update_table_ui(self, data):
        """Update table on main thread"""
        for row in self.table.get_children():
            self.table.delete(row)
        
        for entry in data:
            # Truncate description for display
            description = entry.get('description', '')
            if len(description) > 60:
                description = description[:57] + "..."
            
            self.table.insert("", "end", iid=entry['id'], values=(
                entry['id'],
                entry.get('hs_code', ''),
                entry.get('keyword', ''),
                entry.get('country', ''),
                entry.get('company_name', ''),
                entry.get('company_country', ''),
                entry.get('company_website_link', ''),
                description
            ))
    
    def _clear_table_ui(self):
        """Clear table on main thread"""
        for row in self.table.get_children():
            self.table.delete(row)

    def on_search_change(self, *args):
        """Handle search input changes"""
        self.populate_table()

    def clear_search(self):
        """Clear search and refresh table"""
        self.search_var.set("")
        self.populate_table()

    def load_countries(self):
        """Load available countries from database that have DeepSeek results"""
        try:
            results = GUI_db.get_all_deepseek_results()
            countries_with_results = set()
            for result in results:
                if result.get('company_country'):
                    countries_with_results.add(result['company_country'])
            countries = ["All"] + sorted(list(countries_with_results))
            self.country_list = countries
        except Exception as e:
            print(f"Error loading DeepSeek countries: {e}")
            self.country_list = ["All"]

    def open_country_selector(self):
        """Open a searchable country selection dialog"""
        dialog = CountrySelectorDialog(self, self.country_list)
        self.wait_window(dialog)  # Wait for dialog to close
        selected_country = dialog.get_selected()
        if selected_country:
            self.country_var.set(selected_country)
            # Refresh the table with the new country filter
            self.populate_table()

    def edit_selected(self):
        """Edit the selected record"""
        selected = self.table.selection()
        if not selected:
            messagebox.showwarning("Edit Record", "Please select a row to edit.")
            return
        
        record_id = int(selected[0])
        record = GUI_db.get_deepseek_result_by_id(record_id)
        if not record:
            messagebox.showerror("Not Found", "Selected record not found.")
            return
        
        self.open_edit_dialog(record)

    def open_edit_dialog(self, record):
        """Open edit dialog for a record"""
        dialog = ctk.CTkToplevel(self)
        dialog.title("Edit AI Buyer Result")
        dialog.geometry("800x600")
        dialog.grab_set()
        dialog.transient(self.winfo_toplevel())
        
        ctk.CTkLabel(dialog, text="Edit AI Buyer Result", font=("Poppins", 20, "bold"), text_color="#2E3A59").pack(pady=(20, 16))
        
        # Form fields
        form_frame = ctk.CTkFrame(dialog, fg_color="#F5F7FA")
        form_frame.pack(fill="both", expand=True, padx=10, pady=(0, 8))
        
        # HS Code
        ctk.CTkLabel(form_frame, text="HS Code:", font=("Poppins", 15)).pack(anchor="w", padx=16, pady=(16, 4))
        hs_var = tk.StringVar(value=record.get('hs_code', ''))
        hs_entry = ctk.CTkEntry(form_frame, textvariable=hs_var, font=("Poppins", 14))
        hs_entry.pack(fill="x", padx=16, pady=(0, 12))
        
        # Company Name
        ctk.CTkLabel(form_frame, text="Company Name:", font=("Poppins", 15)).pack(anchor="w", padx=16, pady=(0, 4))
        company_name_var = tk.StringVar(value=record.get('company_name', ''))
        company_name_entry = ctk.CTkEntry(form_frame, textvariable=company_name_var, font=("Poppins", 14))
        company_name_entry.pack(fill="x", padx=16, pady=(0, 12))
        
        # Company Country
        ctk.CTkLabel(form_frame, text="Company Country:", font=("Poppins", 15)).pack(anchor="w", padx=16, pady=(0, 4))
        company_country_var = tk.StringVar(value=record.get('company_country', ''))
        company_country_entry = ctk.CTkEntry(form_frame, textvariable=company_country_var, font=("Poppins", 14))
        company_country_entry.pack(fill="x", padx=16, pady=(0, 12))
        
        # Website
        ctk.CTkLabel(form_frame, text="Website:", font=("Poppins", 15)).pack(anchor="w", padx=16, pady=(0, 4))
        website_var = tk.StringVar(value=record.get('company_website_link', ''))
        website_entry = ctk.CTkEntry(form_frame, textvariable=website_var, font=("Poppins", 14))
        website_entry.pack(fill="x", padx=16, pady=(0, 12))
        
        # Description
        ctk.CTkLabel(form_frame, text="Description:", font=("Poppins", 15)).pack(anchor="w", padx=16, pady=(0, 4))
        desc_border_frame = ctk.CTkFrame(form_frame, fg_color="#B0BEC5", corner_radius=6)
        desc_border_frame.pack(fill="x", padx=16, pady=(0, 12))
        desc_text = ctk.CTkTextbox(desc_border_frame, height=60, font=("Poppins", 14), fg_color="#FFFFFF", border_width=0)
        desc_text.pack(fill="both", expand=True, padx=2, pady=2)
        desc_text.insert("1.0", record.get('description', ''))
        
        # Buttons
        button_frame = ctk.CTkFrame(dialog, fg_color="#F5F7FA")
        button_frame.pack(fill="x", padx=24, pady=(0, 24))
        
        def on_save():
            updated_fields = {
                'hs_code': hs_var.get().strip(),
                'company_name': company_name_var.get().strip(),
                'company_country': company_country_var.get().strip(),
                'company_website_link': website_var.get().strip(),
                'description': desc_text.get("1.0", "end-1c").strip()
            }
            
            if not updated_fields['hs_code'] or not updated_fields['company_name']:
                messagebox.showwarning("Missing Data", "Please fill in HS Code and Company Name.")
                return
            
            if GUI_db.update_deepseek_result(record['id'], updated_fields):
                dialog.destroy()
                self.populate_table()
                messagebox.showinfo("Success", "Record updated successfully.")
            else:
                messagebox.showerror("Error", "Failed to update record.")
        
        ctk.CTkButton(button_frame, text="Save", fg_color="#0078D4", text_color="#FFFFFF", font=("Poppins", 15, "bold"), command=on_save).pack(side="left", padx=(0, 8))
        ctk.CTkButton(button_frame, text="Cancel", fg_color="#B0BEC5", text_color="#FFFFFF", font=("Poppins", 15), command=dialog.destroy).pack(side="right")

    def delete_selected(self):
        """Delete the selected record"""
        selected = self.table.selection()
        if not selected:
            messagebox.showwarning("Delete Record", "Please select a row to delete.")
            return
        
        record_id = int(selected[0])
        if messagebox.askyesno("Delete Record", "Are you sure you want to delete this record?"):
            if GUI_db.delete_deepseek_result(record_id):
                self.populate_table()
                messagebox.showinfo("Success", "Record deleted successfully.")
            else:
                messagebox.showerror("Error", "Failed to delete record.")

    def export_results(self):
        """Export results to CSV"""
        try:
            filename = filedialog.asksaveasfilename(
                defaultextension=".csv",
                filetypes=[("CSV files", "*.csv"), ("All files", "*.*")],
                title="Export AI Buyer Results",
                initialfile="ai_buyer_results.csv"
            )
            
            if filename:
                results = GUI_db.get_all_deepseek_results()
                with open(filename, 'w', newline='', encoding='utf-8') as csvfile:
                    writer = csv.writer(csvfile)
                    # Write headers
                    headers = ["ID", "HS Code", "Keyword", "Company Name", "Company Country", "Website", "Description"]
                    writer.writerow(headers)
                    
                    # Write data
                    for result in results:
                        writer.writerow([
                            result.get('id', ''),
                            result.get('hs_code', ''),
                            result.get('keyword', ''),
                            result.get('company_name', ''),
                            result.get('company_country', ''),
                            result.get('company_website_link', ''),
                            result.get('description', '')
                        ])
                
                messagebox.showinfo("Export Complete", f"Results exported to {filename}")
                
        except Exception as e:
            messagebox.showerror("Export Error", f"Error exporting results: {e}")


class ExportPage(ctk.CTkFrame):
    def __init__(self, master):
        super().__init__(master, fg_color="#F5F7FA")
        # Use a scrollable frame for all content
        self.scrollable = ctk.CTkScrollableFrame(self, fg_color="#F5F7FA")
        self.scrollable.pack(fill="both", expand=True)
        self._build_ui(self.scrollable)
        self.load_data_counts()

    def _build_ui(self, parent):
        # Title
        title_frame = ctk.CTkFrame(parent, fg_color="#F5F7FA")
        title_frame.pack(fill="x", pady=(24, 16), padx=32)
        ctk.CTkLabel(title_frame, text="📊 Data Export Center", font=("Poppins", 24, "bold"), text_color="#2E3A59").pack(side="left")
        
        # Data Overview Cards
        overview_frame = ctk.CTkFrame(parent, fg_color="#F5F7FA")
        overview_frame.pack(fill="x", pady=(0, 16), padx=32)
        overview_frame.grid_columnconfigure((0,1,2,3), weight=1)
        
        # AI Buyer Results Card
        ai_card = ctk.CTkFrame(overview_frame, fg_color="#FFFFFF", corner_radius=12)
        ai_card.grid(row=0, column=0, padx=8, pady=8, sticky="ew")
        ctk.CTkLabel(ai_card, text="🤖 AI Buyer Results", font=("Poppins", 16, "bold"), text_color="#0078D4").pack(pady=(16, 8))
        self.ai_count_label = ctk.CTkLabel(ai_card, text="Loading...", font=("Poppins", 20, "bold"), text_color="#2E3A59")
        self.ai_count_label.pack(pady=(0, 16))
        
        # Apollo Buyer List Card
        apollo_card = ctk.CTkFrame(overview_frame, fg_color="#FFFFFF", corner_radius=12)
        apollo_card.grid(row=0, column=1, padx=8, pady=8, sticky="ew")
        ctk.CTkLabel(apollo_card, text="👥 Apollo Buyer List", font=("Poppins", 16, "bold"), text_color="#4CAF50").pack(pady=(16, 8))
        self.apollo_count_label = ctk.CTkLabel(apollo_card, text="Loading...", font=("Poppins", 20, "bold"), text_color="#2E3A59")
        self.apollo_count_label.pack(pady=(0, 16))
        
        # HS Codes Card
        hs_card = ctk.CTkFrame(overview_frame, fg_color="#FFFFFF", corner_radius=12)
        hs_card.grid(row=0, column=2, padx=8, pady=8, sticky="ew")
        ctk.CTkLabel(hs_card, text="🏷️ HS Codes", font=("Poppins", 16, "bold"), text_color="#FF9800").pack(pady=(16, 8))
        self.hs_count_label = ctk.CTkLabel(hs_card, text="Loading...", font=("Poppins", 20, "bold"), text_color="#2E3A59")
        self.hs_count_label.pack(pady=(0, 16))
        
        # Companies Card
        company_card = ctk.CTkFrame(overview_frame, fg_color="#FFFFFF", corner_radius=12)
        company_card.grid(row=0, column=3, padx=8, pady=8, sticky="ew")
        ctk.CTkLabel(company_card, text="🏢 Companies", font=("Poppins", 16, "bold"), text_color="#9C27B0").pack(pady=(16, 8))
        self.company_count_label = ctk.CTkLabel(company_card, text="Loading...", font=("Poppins", 20, "bold"), text_color="#2E3A59")
        self.company_count_label.pack(pady=(0, 16))
        
        # Export Configuration
        config_frame = ctk.CTkFrame(parent, fg_color="#FFFFFF", corner_radius=16)
        config_frame.pack(fill="x", padx=32, pady=(0, 16))
        
        config_content = ctk.CTkFrame(config_frame, fg_color="transparent")
        config_content.pack(fill="x", padx=20, pady=18)
        
        ctk.CTkLabel(config_content, text="Export Configuration", font=("Poppins", 18, "bold"), text_color="#0078D4").pack(anchor="w", pady=(0, 16))
        
        # Initialize variables for quick export templates
        self.ai_var = tk.BooleanVar(value=False)
        self.apollo_var = tk.BooleanVar(value=False)
        self.hs_var = tk.BooleanVar(value=False)
        self.company_var = tk.BooleanVar(value=False)
        
        # Export Options
        options_frame = ctk.CTkFrame(config_content, fg_color="#F5F7FA", corner_radius=8)
        options_frame.pack(fill="x", pady=(0, 16))
        
        ctk.CTkLabel(options_frame, text="Export Options:", font=("Poppins", 15, "bold"), text_color="#2E3A59").pack(anchor="w", padx=16, pady=(12, 8))
        
        options_grid = ctk.CTkFrame(options_frame, fg_color="transparent")
        options_grid.pack(fill="x", padx=16, pady=(0, 12))
        options_grid.grid_columnconfigure((0,1,2), weight=1)
        
        # Format Selection
        ctk.CTkLabel(options_grid, text="Format:", font=("Poppins", 14)).grid(row=0, column=0, sticky="w", padx=8, pady=4)
        self.format_var = tk.StringVar(value="CSV")
        format_combo = ctk.CTkComboBox(options_grid, variable=self.format_var, values=["CSV", "Excel"], font=("Poppins", 14))
        format_combo.grid(row=0, column=1, sticky="w", padx=8, pady=4)
        
        # Include Headers
        self.headers_var = tk.BooleanVar(value=True)
        ctk.CTkCheckBox(options_grid, text="Include Headers", variable=self.headers_var, font=("Poppins", 14)).grid(row=0, column=2, sticky="w", padx=8, pady=4)
        
        # Initialize filter variables (not used in simplified version)
        self.country_var = tk.StringVar(value="All")
        self.date_var = tk.StringVar(value="All Time")
        
        # Action Buttons
        button_frame = ctk.CTkFrame(config_content, fg_color="transparent")
        button_frame.pack(fill="x", pady=(16, 0))
        
        # Quick Export Templates
        templates_frame = ctk.CTkFrame(button_frame, fg_color="#F5F7FA", corner_radius=8)
        templates_frame.pack(fill="x", pady=(0, 16))
        
        ctk.CTkLabel(templates_frame, text="Quick Export Templates:", font=("Poppins", 15, "bold"), text_color="#2E3A59").pack(anchor="w", padx=16, pady=(12, 8))
        
        templates_grid = ctk.CTkFrame(templates_frame, fg_color="transparent")
        templates_grid.pack(fill="x", padx=16, pady=(0, 12))
        templates_grid.grid_columnconfigure((0,1,2), weight=1)
        
        ctk.CTkButton(templates_grid, text="🤖 AI Buyer Results", fg_color="#4CAF50", hover_color="#388E3C", text_color="#FFFFFF", font=("Poppins", 12, "bold"), command=self.export_sales_report).grid(row=0, column=0, padx=8, pady=4, sticky="ew")
        ctk.CTkButton(templates_grid, text="📋 HS Codes", fg_color="#FF9800", hover_color="#F57C00", text_color="#FFFFFF", font=("Poppins", 12, "bold"), command=self.export_market_analysis).grid(row=0, column=1, padx=8, pady=4, sticky="ew")
        ctk.CTkButton(templates_grid, text="👥 Apollo Buyer List", fg_color="#9C27B0", hover_color="#7B1FA2", text_color="#FFFFFF", font=("Poppins", 12, "bold"), command=self.export_lead_list).grid(row=0, column=2, padx=8, pady=4, sticky="ew")
        

        
        # Progress Bar
        self.progress_bar = ctk.CTkProgressBar(button_frame, orientation="horizontal", mode="determinate", width=400)
        self.progress_bar.pack(pady=8)
        self.progress_bar.set(0)
        self.progress_bar.pack_forget()
        
        # Status Label
        self.status_label = ctk.CTkLabel(button_frame, text="", font=("Poppins", 14), text_color="#6B7C93")
        self.status_label.pack(pady=4)

    def load_data_counts(self):
        """Load and display data counts for each source"""
        try:
            # AI Buyer Results count
            ai_count = len(GUI_db.get_all_deepseek_results())
            
            # Apollo Buyer List count
            apollo_count = len(GUI_db.get_all_contacts())
            
            # HS Codes count
            hs_count = len(GUI_db.get_all_hs_codes())
            
            # Companies count
            company_count = len(GUI_db.get_all_companies())
            
            # Get unique countries for filter
            countries = set()
            for result in GUI_db.get_all_deepseek_results():
                if result.get('company_country'):
                    countries.add(result['company_country'])
            for contact in GUI_db.get_all_contacts():
                if contact.get('company_name'):
                    # Extract country from company name or use a default
                    countries.add("Unknown")
            for hs_code in GUI_db.get_all_hs_codes():
                if hs_code.get('country'):
                    countries.add(hs_code['country'])
            
            counts = {
                'ai_count': ai_count,
                'apollo_count': apollo_count,
                'hs_count': hs_count,
                'company_count': company_count,
                'countries': sorted(list(countries))
            }
            
            # Update UI directly on main thread
            self._update_counts_ui(counts)
            
        except Exception as error:
            print(f"Error loading data counts: {error}")
            self._update_counts_ui({
                'ai_count': 0,
                'apollo_count': 0,
                'hs_count': 0,
                'company_count': 0,
                'countries': ["All"]
            })

    def _update_counts_ui(self, counts):
        """Update the count labels on main thread"""
        self.ai_count_label.configure(text=str(counts['ai_count']))
        self.apollo_count_label.configure(text=str(counts['apollo_count']))
        self.hs_count_label.configure(text=str(counts['hs_count']))
        self.company_count_label.configure(text=str(counts['company_count']))

    def export_sales_report(self):
        """Export AI Buyer Results table"""
        self.ai_var.set(True)
        self.apollo_var.set(False)
        self.company_var.set(False)
        self.hs_var.set(False)
        self.format_var.set("Excel")
        self.export_data()

    def export_market_analysis(self):
        """Export HS Codes table"""
        self.ai_var.set(False)
        self.hs_var.set(True)
        self.company_var.set(False)
        self.apollo_var.set(False)
        self.format_var.set("Excel")
        self.export_data()

    def export_lead_list(self):
        """Export Apollo Buyer List table"""
        self.apollo_var.set(True)
        self.ai_var.set(False)
        self.company_var.set(False)
        self.hs_var.set(False)
        self.format_var.set("CSV")
        self.export_data()

    def export_data(self):
        """Main export function"""
        # Check if any data source is selected
        if not any([self.ai_var.get(), self.apollo_var.get(), self.hs_var.get(), self.company_var.get()]):
            messagebox.showwarning("No Data Selected", "Please use one of the quick export templates.")
            return
        
        # Show progress bar
        self.progress_bar.pack(pady=8)
        self.progress_bar.set(0)
        self.status_label.configure(text="Preparing export...")
        
        def run_export():
            try:
                import csv
                from datetime import datetime, timedelta
                
                # Get filter values
                selected_country = self.country_var.get()
                date_filter = self.date_var.get()
                export_format = self.format_var.get()
                include_headers = self.headers_var.get()
                
                # Calculate date range
                end_date = datetime.now()
                if date_filter == "Last 7 Days":
                    start_date = end_date - timedelta(days=7)
                elif date_filter == "Last 30 Days":
                    start_date = end_date - timedelta(days=30)
                elif date_filter == "Last 90 Days":
                    start_date = end_date - timedelta(days=90)
                else:
                    start_date = None
                
                # Generate filename
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                if export_format == "CSV":
                    filename = f"buyer_intelligence_export_{timestamp}.csv"
                else:
                    filename = f"buyer_intelligence_export_{timestamp}.xlsx"
                
                # Get file path
                file_path = filedialog.asksaveasfilename(
                    defaultextension=f".{export_format.lower()}",
                    filetypes=[(f"{export_format} files", f"*.{export_format.lower()}"), ("All files", "*.*")],
                    title="Save Export File",
                    initialfile=filename
                )
                
                if not file_path:
                    return
                
                # Collect and enrich data based on selections
                all_data = {}
                
                # Load all data first for cross-referencing
                ai_data = []
                apollo_data = []
                company_data = []
                hs_data = []
                
                if self.ai_var.get():
                    self.after(0, lambda: self.status_label.configure(text="Loading AI Buyer Results..."))
                    self.after(0, lambda: self.progress_bar.set(0.1))
                    
                    ai_data = GUI_db.get_all_deepseek_results()
                    if selected_country != "All":
                        ai_data = [r for r in ai_data if r.get('company_country') == selected_country]
                    if start_date:
                        ai_data = [r for r in ai_data if self._parse_date(r.get('created_at', '')) >= start_date]
                
                if self.apollo_var.get():
                    self.after(0, lambda: self.status_label.configure(text="Loading Apollo Buyer List..."))
                    self.after(0, lambda: self.progress_bar.set(0.2))
                    
                    apollo_data = GUI_db.get_all_contacts()
                    if selected_country != "All":
                        apollo_data = [c for c in apollo_data if self._get_contact_country(c) == selected_country]
                    if start_date:
                        apollo_data = [c for c in apollo_data if self._parse_date(c.get('created_at', '')) >= start_date]
                
                if self.hs_var.get():
                    self.after(0, lambda: self.status_label.configure(text="Loading HS Codes..."))
                    self.after(0, lambda: self.progress_bar.set(0.3))
                    
                    hs_data = GUI_db.get_all_hs_codes()
                    if selected_country != "All":
                        hs_data = [h for h in hs_data if h.get('country') == selected_country]
                    if start_date:
                        hs_data = [h for h in hs_data if self._parse_date(h.get('created_at', '')) >= start_date]
                
                if self.company_var.get():
                    self.after(0, lambda: self.status_label.configure(text="Loading Companies..."))
                    self.after(0, lambda: self.progress_bar.set(0.4))
                    
                    company_data = GUI_db.get_all_companies()
                    if selected_country != "All":
                        company_data = [c for c in company_data if c.get('country') == selected_country]
                    if start_date:
                        company_data = [c for c in company_data if self._parse_date(c.get('created_at', '')) >= start_date]
                
                # Export raw data tables
                self.after(0, lambda: self.status_label.configure(text="Preparing export..."))
                self.after(0, lambda: self.progress_bar.set(0.5))
                
                # Export selected tables
                if self.ai_var.get():
                    all_data['AI Buyer Results'] = ai_data
                if self.apollo_var.get():
                    all_data['Apollo Buyer List'] = apollo_data
                if self.hs_var.get():
                    all_data['HS Codes'] = hs_data
                if self.company_var.get():
                    all_data['Companies'] = company_data
                
                # Export based on format
                self.after(0, lambda: self.status_label.configure(text="Exporting data..."))
                self.after(0, lambda: self.progress_bar.set(0.9))
                
                if export_format == "CSV":
                    self._export_to_csv(file_path, all_data, include_headers)
                else:
                    self._export_to_excel(file_path, all_data, include_headers)
                
                self.after(0, lambda: self.progress_bar.set(1.0))
                self.after(0, lambda: self.status_label.configure(text="Export completed successfully!"))
                self.after(0, lambda: messagebox.showinfo("Export Complete", f"Data exported successfully to:\n{file_path}"))
                
            except Exception as e:
                self.after(0, lambda: self.status_label.configure(text=f"Export failed: {str(e)}"))
                self.after(0, lambda: messagebox.showerror("Export Error", f"Error during export: {str(e)}"))
            finally:
                self.after(0, lambda: self.progress_bar.pack_forget())
                self.after(0, lambda: self.status_label.configure(text=""))
        
        threading.Thread(target=run_export, daemon=True).start()

    def _parse_date(self, date_str):
        """Parse date string to datetime object"""
        if not date_str:
            return datetime.min
        try:
            # Try different date formats
            for fmt in ['%Y-%m-%dT%H:%M:%S.%f', '%Y-%m-%dT%H:%M:%S', '%Y-%m-%d %H:%M:%S', '%Y-%m-%d']:
                try:
                    return datetime.strptime(date_str, fmt)
                except ValueError:
                    continue
            return datetime.min
        except:
            return datetime.min

    def _get_contact_country(self, contact):
        """Get country for a contact (placeholder implementation)"""
        # This would need to be implemented based on your data structure
        return contact.get('country', 'Unknown')

    def _create_sales_report(self, ai_data, apollo_data, company_data, hs_data):
        """Create an enhanced sales report with contact information and company insights"""
        sales_report = []
        
        # Create company lookup
        company_lookup = {c['company_name'].lower(): c for c in company_data if c.get('company_name')}
        
        # Create HS code lookup
        hs_lookup = {h['hs_code']: h['description'] for h in hs_data if h.get('hs_code')}
        
        # Group contacts by company
        contacts_by_company = defaultdict(list)
        for contact in apollo_data:
            company_name = contact.get('company_name', '').lower()
            if company_name:
                contacts_by_company[company_name].append(contact)
        
        # Create sales report entries
        for ai_company in ai_data:
            company_name = ai_company.get('company_name', '')
            company_name_lower = company_name.lower()
            
            # Get company info
            company_info = company_lookup.get(company_name_lower, {})
            
            # Get contacts for this company
            contacts = contacts_by_company.get(company_name_lower, [])
            
            # Get HS code description
            hs_code = ai_company.get('hs_code', '')
            hs_description = hs_lookup.get(hs_code, '')
            
            # Create sales report entry
            entry = {
                'Company Name': company_name,
                'Company Country': ai_company.get('company_country', ''),
                'Company Website': ai_company.get('company_website_link', ''),
                'Company Industry': company_info.get('industry', ''),
                'Company Size': company_info.get('employee_count', ''),
                'HS Code': hs_code,
                'HS Description': hs_description,
                'Product Keyword': ai_company.get('keyword', ''),
                'Search Country': ai_company.get('country', ''),
                'Company Description': ai_company.get('description', ''),
                'Contact Count': len(contacts),
                'Primary Contact Name': contacts[0].get('name', '') if contacts else '',
                'Primary Contact Title': contacts[0].get('title', '') if contacts else '',
                'Primary Contact Email': contacts[0].get('email', '') if contacts else '',
                'Primary Contact LinkedIn': contacts[0].get('linkedin', '') if contacts else '',
                'All Contact Emails': '; '.join([c.get('email', '') for c in contacts if c.get('email')]),
                'All Contact LinkedIn': '; '.join([c.get('linkedin', '') for c in contacts if c.get('linkedin')]),
                'Lead Score': self._calculate_lead_score(company_info, contacts, ai_company),
                'Source': ai_company.get('source', ''),
                'Created Date': ai_company.get('created_at', '')
            }
            sales_report.append(entry)
        
        return sales_report

    def _create_market_analysis(self, ai_data, company_data, hs_data):
        """Create a market analysis report with trends and insights"""
        market_analysis = []
        
        # Analyze market trends
        hs_code_counts = Counter()
        country_counts = Counter()
        keyword_counts = Counter()
        
        for company in ai_data:
            hs_code_counts[company.get('hs_code', '')] += 1
            country_counts[company.get('company_country', '')] += 1
            keyword_counts[company.get('keyword', '')] += 1
        
        # Create market analysis entries
        for hs_code, count in hs_code_counts.most_common():
            hs_description = next((h['description'] for h in hs_data if h.get('hs_code') == hs_code), '')
            
            # Get companies for this HS code
            companies_for_hs = [c for c in ai_data if c.get('hs_code') == hs_code]
            countries_for_hs = Counter(c.get('company_country', '') for c in companies_for_hs)
            
            entry = {
                'HS Code': hs_code,
                'HS Description': hs_description,
                'Total Companies': count,
                'Top Countries': '; '.join([f"{country} ({count})" for country, count in countries_for_hs.most_common(5)]),
                'Market Share %': round((count / len(ai_data)) * 100, 2) if ai_data else 0,
                'Average Company Size': self._calculate_avg_company_size(companies_for_hs, company_data),
                'Keywords Used': '; '.join(set(c.get('keyword', '') for c in companies_for_hs)),
                'Market Trend': self._determine_market_trend(companies_for_hs),
                'Source': 'DeepSeek Analysis'
            }
            market_analysis.append(entry)
        
        return market_analysis

    def _create_lead_scoring(self, apollo_data, ai_data, company_data):
        """Create a lead scoring report with prioritization"""
        lead_scoring = []
        
        # Create company lookup
        company_lookup = {c['company_name'].lower(): c for c in company_data if c.get('company_name')}
        
        # Group contacts by company
        contacts_by_company = defaultdict(list)
        for contact in apollo_data:
            company_name = contact.get('company_name', '').lower()
            if company_name:
                contacts_by_company[company_name].append(contact)
        
        # Score each company
        for company_name, contacts in contacts_by_company.items():
            company_info = company_lookup.get(company_name, {})
            
            # Find AI data for this company
            ai_company = next((c for c in ai_data if c.get('company_name', '').lower() == company_name), {})
            
            # Calculate lead score
            lead_score = self._calculate_lead_score(company_info, contacts, ai_company)
            
            # Determine priority
            if lead_score >= 80:
                priority = "High"
            elif lead_score >= 60:
                priority = "Medium"
            else:
                priority = "Low"
            
            entry = {
                'Company Name': contacts[0].get('company_name', '') if contacts else company_name,
                'Lead Score': lead_score,
                'Priority': priority,
                'Contact Count': len(contacts),
                'Has Email': any(c.get('email') for c in contacts),
                'Has LinkedIn': any(c.get('linkedin') for c in contacts),
                'Company Size': company_info.get('employee_count', ''),
                'Company Industry': company_info.get('industry', ''),
                'Company Country': company_info.get('country', ''),
                'HS Code Interest': ai_company.get('hs_code', ''),
                'Product Interest': ai_company.get('keyword', ''),
                'Primary Contact': contacts[0].get('name', '') if contacts else '',
                'Primary Contact Title': contacts[0].get('title', '') if contacts else '',
                'Primary Contact Email': contacts[0].get('email', '') if contacts else '',
                'All Emails': '; '.join([c.get('email', '') for c in contacts if c.get('email')]),
                'All LinkedIn': '; '.join([c.get('linkedin', '') for c in contacts if c.get('linkedin')]),
                'Source': contacts[0].get('source', '') if contacts else '',
                'Created Date': contacts[0].get('created_at', '') if contacts else ''
            }
            lead_scoring.append(entry)
        
        # Sort by lead score descending
        lead_scoring.sort(key=lambda x: x['Lead Score'], reverse=True)
        return lead_scoring

    def _create_company_intelligence(self, company_data, apollo_data, ai_data):
        """Create a company intelligence report with comprehensive insights"""
        company_intelligence = []
        
        # Create contact counts
        contact_counts = Counter()
        for contact in apollo_data:
            company_name = contact.get('company_name', '').lower()
            if company_name:
                contact_counts[company_name] += 1
        
        # Create AI interest counts
        ai_interest_counts = Counter()
        for company in ai_data:
            company_name = company.get('company_name', '').lower()
            if company_name:
                ai_interest_counts[company_name] += 1
        
        for company in company_data:
            company_name = company.get('company_name', '')
            company_name_lower = company_name.lower()
            
            # Get contact count
            contact_count = contact_counts.get(company_name_lower, 0)
            
            # Get AI interest count
            ai_interest_count = ai_interest_counts.get(company_name_lower, 0)
            
            # Get contacts for this company
            contacts = [c for c in apollo_data if c.get('company_name', '').lower() == company_name_lower]
            
            # Get AI interests for this company
            ai_interests = [c for c in ai_data if c.get('company_name', '').lower() == company_name_lower]
            
            entry = {
                'Company Name': company_name,
                'Company Country': company.get('country', ''),
                'Company Industry': company.get('industry', ''),
                'Company Size': company.get('employee_count', ''),
                'Company Domain': company.get('domain', ''),
                'Contact Count': contact_count,
                'AI Interest Count': ai_interest_count,
                'Total Engagement Score': contact_count + ai_interest_count,
                'Has Verified Contacts': contact_count > 0,
                'Has AI Interest': ai_interest_count > 0,
                'Contact Emails': '; '.join(set(c.get('email', '') for c in contacts if c.get('email'))),
                'Contact LinkedIn': '; '.join(set(c.get('linkedin', '') for c in contacts if c.get('linkedin'))),
                'Interested HS Codes': '; '.join(set(c.get('hs_code', '') for c in ai_interests if c.get('hs_code'))),
                'Interested Keywords': '; '.join(set(c.get('keyword', '') for c in ai_interests if c.get('keyword'))),
                'Company Intelligence Score': self._calculate_company_intelligence_score(company, contacts, ai_interests),
                'Source': company.get('source', ''),
                'Created Date': company.get('created_at', '')
            }
            company_intelligence.append(entry)
        
        # Sort by intelligence score descending
        company_intelligence.sort(key=lambda x: x['Company Intelligence Score'], reverse=True)
        return company_intelligence

    def _calculate_lead_score(self, company_info, contacts, ai_company):
        """Calculate a lead score based on various factors"""
        score = 0
        
        # Company size factor (0-20 points)
        employee_count = company_info.get('employee_count', 0)
        if employee_count:
            if employee_count > 1000:
                score += 20
            elif employee_count > 500:
                score += 15
            elif employee_count > 100:
                score += 10
            elif employee_count > 50:
                score += 5
        
        # Contact quality factor (0-30 points)
        if contacts:
            score += min(len(contacts) * 5, 15)  # Up to 15 points for contact count
            if any(c.get('email') for c in contacts):
                score += 10  # 10 points for having email
            if any(c.get('linkedin') for c in contacts):
                score += 5   # 5 points for having LinkedIn
        
        # AI interest factor (0-30 points)
        if ai_company:
            score += 20  # Base points for AI interest
            if ai_company.get('company_website_link'):
                score += 5   # 5 points for having website
            if ai_company.get('description'):
                score += 5   # 5 points for having description
        
        # Industry factor (0-20 points)
        industry = company_info.get('industry', '').lower()
        if any(keyword in industry for keyword in ['medical', 'healthcare', 'hospital', 'clinic', 'pharmaceutical']):
            score += 20
        elif any(keyword in industry for keyword in ['manufacturing', 'industrial', 'construction', 'automotive']):
            score += 15
        elif any(keyword in industry for keyword in ['retail', 'wholesale', 'distribution', 'logistics']):
            score += 10
        
        return min(score, 100)  # Cap at 100

    def _calculate_avg_company_size(self, companies, company_data):
        """Calculate average company size for a group of companies"""
        company_lookup = {c['company_name'].lower(): c for c in company_data if c.get('company_name')}
        
        sizes = []
        for company in companies:
            company_name = company.get('company_name', '').lower()
            company_info = company_lookup.get(company_name, {})
            employee_count = company_info.get('employee_count', 0)
            if employee_count:
                sizes.append(employee_count)
        
        if sizes:
            return sum(sizes) // len(sizes)
        return 0

    def _determine_market_trend(self, companies):
        """Determine market trend based on company data"""
        if not companies:
            return "No Data"
        
        # Simple trend determination based on company count
        if len(companies) > 10:
            return "High Growth"
        elif len(companies) > 5:
            return "Growing"
        elif len(companies) > 2:
            return "Stable"
        else:
            return "Emerging"

    def _calculate_company_intelligence_score(self, company, contacts, ai_interests):
        """Calculate company intelligence score"""
        score = 0
        
        # Base company info (0-20 points)
        if company.get('industry'):
            score += 10
        if company.get('employee_count'):
            score += 10
        
        # Contact quality (0-30 points)
        if contacts:
            score += min(len(contacts) * 3, 15)
            if any(c.get('email') for c in contacts):
                score += 10
            if any(c.get('linkedin') for c in contacts):
                score += 5
        
        # AI interest (0-30 points)
        if ai_interests:
            score += min(len(ai_interests) * 5, 20)
            if any(c.get('company_website_link') for c in ai_interests):
                score += 5
            if any(c.get('description') for c in ai_interests):
                score += 5
        
        # Engagement factor (0-20 points)
        total_engagement = len(contacts) + len(ai_interests)
        if total_engagement > 5:
            score += 20
        elif total_engagement > 3:
            score += 15
        elif total_engagement > 1:
            score += 10
        
        return min(score, 100)

    def _export_to_csv(self, file_path, all_data, include_headers):
        """Export data to CSV format"""
        import csv
        
        with open(file_path, 'w', newline='', encoding='utf-8') as csvfile:
            writer = csv.writer(csvfile)
            
            for sheet_name, data in all_data.items():
                if not data:
                    continue
                
                # Write sheet name as section header
                writer.writerow([f"=== {sheet_name} ==="])
                
                if include_headers and data:
                    # Write headers
                    headers = list(data[0].keys())
                    writer.writerow(headers)
                
                # Write data
                for row in data:
                    writer.writerow([row.get(header, '') for header in headers])
                
                # Add spacing between sections
                writer.writerow([])
                writer.writerow([])

    def _export_to_excel(self, file_path, all_data, include_headers):
        """Export data to Excel format"""
        try:
            import pandas as pd
            from openpyxl import Workbook
            from openpyxl.utils.dataframe import dataframe_to_rows
            
            # Create Excel workbook
            wb = Workbook()
            
            # Remove default sheet if it exists
            if wb.active is not None:
                wb.remove(wb.active)
            
            for sheet_name, data in all_data.items():
                if not data:
                    continue
                
                # Create DataFrame
                df = pd.DataFrame(data)
                
                # Create worksheet
                ws = wb.create_sheet(title=sheet_name[:31])  # Excel sheet names limited to 31 chars
                
                # Write headers
                if include_headers:
                    for col, header in enumerate(df.columns, 1):
                        ws.cell(row=1, column=col, value=header)
                
                # Write data
                for row_idx, row in enumerate(dataframe_to_rows(df, index=False, header=False), 2 if include_headers else 1):
                    for col_idx, value in enumerate(row, 1):
                        ws.cell(row=row_idx, column=col_idx, value=value)
            
            # Save workbook
            wb.save(file_path)
            
        except ImportError:
            # Fallback to CSV if pandas/openpyxl not available
            messagebox.showwarning("Excel Export", "Pandas/OpenPyXL not available. Falling back to CSV format.")
            self._export_to_csv(file_path, all_data, include_headers)

class HSCodePage(ctk.CTkFrame):
    def __init__(self, master):
        super().__init__(master, fg_color="#F5F7FA")
        self._build_ui()
        # Populate table by default
        self.populate_table()

    def _build_ui(self):
        # Top bar: Add HS Code button (right)
        topbar = ctk.CTkFrame(self, fg_color="#F5F7FA")
        topbar.pack(fill="x", pady=(24, 0), padx=32)
        ctk.CTkLabel(topbar, text="HS Code Manager", font=("Poppins", 22, "bold"), text_color="#2E3A59").pack(side="left")
        add_btn = ctk.CTkButton(topbar, text="Add HS Code", fg_color="#0078D4", hover_color="#005A9E", text_color="#FFFFFF", font=("Poppins", 15, "bold"), corner_radius=8, command=self.add_hs_code)
        add_btn.pack(side="right")

        # Search/filter row
        filter_frame = ctk.CTkFrame(self, fg_color="#F5F7FA")
        filter_frame.pack(fill="x", pady=(18, 0), padx=32)
        ctk.CTkLabel(filter_frame, text="Country:", font=("Poppins", 15), text_color="#6B7C93").pack(side="left", padx=(0, 8))
        self.country_var = tk.StringVar(value="All")
        country_frame = ctk.CTkFrame(filter_frame, fg_color="transparent")
        country_frame.pack(side="left", padx=(0, 16))
        
        # Country display (read-only)
        self.country_display = ctk.CTkEntry(country_frame, textvariable=self.country_var, state="readonly", width=160, font=("Poppins", 15))
        self.country_display.pack(side="left")
        
        # Select country button
        select_country_btn = ctk.CTkButton(country_frame, text="Select", fg_color="#4CAF50", hover_color="#388E3C", text_color="#FFFFFF", font=("Poppins", 12, "bold"), width=60, height=32, command=self.open_country_selector)
        select_country_btn.pack(side="left", padx=(8, 0))
        ctk.CTkLabel(filter_frame, text="HS Code:", font=("Poppins", 15), text_color="#6B7C93").pack(side="left", padx=(16, 8))
        self.search_var = tk.StringVar()
        search_entry = ctk.CTkEntry(filter_frame, textvariable=self.search_var, placeholder_text="Enter HS code or description...", width=220, font=("Poppins", 15))
        search_entry.pack(side="left", padx=(0, 12))
        search_btn = ctk.CTkButton(filter_frame, text="Search", fg_color="#0078D4", hover_color="#005A9E", text_color="#FFFFFF", font=("Poppins", 15, "bold"), corner_radius=8, width=90, command=self.do_search)
        search_btn.pack(side="left", padx=(0, 8))
        refresh_btn = ctk.CTkButton(filter_frame, text="Refresh Table", fg_color="#4CAF50", hover_color="#388E3C", text_color="#FFFFFF", font=("Poppins", 15), corner_radius=8, width=120, command=self.do_search)
        refresh_btn.pack(side="left")

        # Table frame
        table_frame = ctk.CTkFrame(self, fg_color="#FFFFFF", corner_radius=16)
        table_frame.pack(fill="both", expand=True, padx=32, pady=(18, 0))
        style = ttk.Style()
        style.configure("Zen.Treeview", font=("Poppins", 13), rowheight=32, background="#FFFFFF", fieldbackground="#FFFFFF", foreground="#2E3A59")
        style.configure("Zen.Treeview.Heading", font=("Poppins", 14, "bold"), background="#F5F7FA", foreground="#0078D4")
        style.map("Zen.Treeview", background=[("selected", "#1E3A8A")], foreground=[("selected", "#FFFFFF")])
        self.table = ttk.Treeview(table_frame, columns=("country", "hs_code", "desc"), show="headings", style="Zen.Treeview")
        self.table.heading("country", text="Country")
        self.table.heading("hs_code", text="HS Code")
        self.table.heading("desc", text="Description")
        self.table.column("country", width=120, anchor="center")
        self.table.column("hs_code", width=120, anchor="center")
        self.table.column("desc", width=400, anchor="w")
        self.table.pack(fill="both", expand=True, padx=8, pady=8)
        self._add_table_sorting()

        # Action buttons below table
        action_frame = ctk.CTkFrame(self, fg_color="#F5F7FA")
        action_frame.pack(fill="x", padx=32, pady=(8, 24))
        edit_btn = ctk.CTkButton(action_frame, text="Edit Selected", fg_color="#4CAF50", hover_color="#388E3C", text_color="#FFFFFF", font=("Poppins", 15), corner_radius=8, width=140, command=self.edit_selected)
        delete_btn = ctk.CTkButton(action_frame, text="Delete Selected", fg_color="#B0BEC5", hover_color="#B0BEC5", text_color="#FFFFFF", font=("Poppins", 15), corner_radius=8, width=140, command=self.delete_selected)
        edit_btn.pack(side="right", padx=(0, 12))
        delete_btn.pack(side="right", padx=(0, 12))

    def populate_table(self):
        try:
            country = self.country_var.get()
            query = self.search_var.get().strip().lower()
            if country == "All":
                data = GUI_db.get_all_hs_codes()
            else:
                data = GUI_db.get_hs_codes_by_country(country)
            # Filter data
            filtered_data = []
            for entry in data:
                if query and query not in entry['hs_code'].lower() and query not in entry['description'].lower():
                    continue
                filtered_data.append(entry)
            # Sort by country and hs_code ascending
            filtered_data = sorted(filtered_data, key=lambda x: (x.get('country', ''), x.get('hs_code', '')))
            
            # Update UI directly on main thread
            self._update_table_ui(filtered_data)
            
        except Exception as error:
            print(f"Error loading table data: {error}")
            self._clear_table_ui()

    def _add_table_sorting(self):
        for col in self.table['columns']:
            self.table.heading(col, command=lambda c=col: self._sort_by_column(c, False))
    def _sort_by_column(self, col, descending):
        data = [(self.table.set(child, col), child) for child in self.table.get_children('')]
        data.sort(reverse=descending)
        for index, (val, child) in enumerate(data):
            self.table.move(child, '', index)
        self.table.heading(col, command=lambda: self._sort_by_column(col, not descending))
    def _update_table_ui(self, data):
        """Update table on main thread"""
        # Clear existing rows
        for row in self.table.get_children():
            self.table.delete(row)
        
        # Insert new data
        for entry in data:
            self.table.insert("", "end", iid=entry['id'], values=(entry['country'], entry['hs_code'], entry['description']))
    
    def _clear_table_ui(self):
        """Clear table on main thread"""
        for row in self.table.get_children():
            self.table.delete(row)

    def do_search(self):
        self.populate_table()

    def open_country_selector(self):
        """Open a searchable country selection dialog"""
        # Get countries that have HS codes in the database
        try:
            hs_codes = GUI_db.get_all_hs_codes()
            countries_with_hs_codes = set()
            for code in hs_codes:
                if code.get('country'):
                    countries_with_hs_codes.add(code['country'])
            
            # Sort the countries and add "All" option
            country_list = ["All"] + sorted(list(countries_with_hs_codes))
        except Exception as e:
            print(f"Error loading countries for HS Code Manager: {e}")
            country_list = ["All"]
        
        dialog = CountrySelectorDialog(self, country_list)
        self.wait_window(dialog)  # Wait for dialog to close
        selected_country = dialog.get_selected()
        if selected_country:
            self.country_var.set(selected_country)
            # Refresh the table with the new country filter
            self.populate_table()

    def open_add_country_selector(self, dialog, country_var):
        """Open a searchable country selection dialog for Add HS Code"""
        country_list = ["Select Country"] + GUI_db.get_all_available_countries()
        selector_dialog = CountrySelectorDialog(dialog, country_list)
        dialog.wait_window(selector_dialog)  # Wait for dialog to close
        selected_country = selector_dialog.get_selected()
        if selected_country and selected_country != "Select Country":
            country_var.set(selected_country)

    def add_hs_code(self):
        dialog = ctk.CTkToplevel(self)
        dialog.title("Add HS Code")
        dialog.geometry("500x600")  # Increased height
        dialog.grab_set()
        
        ctk.CTkLabel(dialog, text="Add HS Code", font=("Poppins", 20, "bold"), text_color="#2E3A59").pack(pady=(24, 16))
        
        # Country selection
        country_frame = ctk.CTkFrame(dialog, fg_color="#F5F7FA")
        country_frame.pack(fill="x", padx=24, pady=8)
        ctk.CTkLabel(country_frame, text="Country:", font=("Poppins", 15)).pack(anchor="w", pady=(12, 4))
        country_var = tk.StringVar(value="Select Country")
        country_display_frame = ctk.CTkFrame(country_frame, fg_color="transparent")
        country_display_frame.pack(fill="x", pady=4)
        country_display = ctk.CTkEntry(country_display_frame, textvariable=country_var, state="readonly", font=("Poppins", 14))
        country_display.pack(side="left", fill="x", expand=True)
        select_country_btn = ctk.CTkButton(country_display_frame, text="Select", fg_color="#4CAF50", hover_color="#388E3C", text_color="#FFFFFF", font=("Poppins", 12, "bold"), width=60, height=32, command=lambda: self.open_add_country_selector(dialog, country_var))
        select_country_btn.pack(side="right", padx=(8, 0))
        
        # HS Code input
        hs_frame = ctk.CTkFrame(dialog, fg_color="#F5F7FA")
        hs_frame.pack(fill="x", padx=24, pady=8)
        ctk.CTkLabel(hs_frame, text="HS Code:", font=("Poppins", 15)).pack(anchor="w", pady=(12, 4))
        hs_var = tk.StringVar()
        hs_entry = ctk.CTkEntry(hs_frame, textvariable=hs_var, placeholder_text="Enter HS Code", font=("Poppins", 14))
        hs_entry.pack(fill="x", pady=4)
        
        # Description input
        desc_frame = ctk.CTkFrame(dialog, fg_color="#F5F7FA")
        desc_frame.pack(fill="x", padx=24, pady=8)
        ctk.CTkLabel(desc_frame, text="Description:", font=("Poppins", 15)).pack(anchor="w", pady=(12, 4))
        desc_text = ctk.CTkTextbox(desc_frame, height=100, font=("Poppins", 14))
        desc_text.pack(fill="x", pady=4)
        
        # Progress label
        progress_label = ctk.CTkLabel(dialog, text="", font=("Poppins", 12), text_color="#6B7C93")
        progress_label.pack(pady=(8, 0))
        # Progress bar
        progress_bar = ctk.CTkProgressBar(dialog, orientation="horizontal", mode="indeterminate", width=320)
        progress_bar.pack(pady=(4, 8))
        progress_bar.set(0)
        progress_bar.pack_forget()  # Hide initially
        
        # Buttons
        button_frame = ctk.CTkFrame(dialog, fg_color="#F5F7FA")
        button_frame.pack(fill="x", padx=24, pady=(16, 24))
        
        def on_manual_save():
            hs = hs_var.get().strip()
            desc = desc_text.get("1.0", "end-1c").strip()
            country = country_var.get().strip()
            if not hs or not desc or not country:
                messagebox.showwarning("Missing Data", "Please fill in all fields.")
                return
            if GUI_db.save_hs_code(hs, desc, country):
                dialog.destroy()
                self.populate_table()
                
                # Refresh other pages that depend on HS codes
                self.refresh_other_pages()
            else:
                messagebox.showerror("Duplicate", "This HS code for the selected country already exists.")
        
        def on_deepseek_search():
            country = country_var.get().strip()
            if not country:
                messagebox.showwarning("Missing Country", "Please select a country to search for HS codes.")
                return
            progress_label.configure(text="Searching DeepSeek for HS codes...")
            progress_bar.pack(pady=(4, 8))
            progress_bar.start()
            dialog.update()
            
            def run_deepseek():
                try:
                    raw_response = deepseek_agent.query_deepseek_for_hs_codes(country)
                    parsed_codes = deepseek_agent.parse_hs_codes_from_deepseek(raw_response)
                except Exception as e:
                    def on_error():
                        progress_bar.stop()
                        progress_bar.pack_forget()
                        progress_label.configure(text="")
                        messagebox.showerror("DeepSeek Error", f"Error calling DeepSeek: {str(e)}")
                    dialog.after(0, on_error)
                    return
                def on_done():
                    progress_bar.stop()
                    progress_bar.pack_forget()
                    if not parsed_codes:
                        progress_label.configure(text="")
                        messagebox.showinfo("DeepSeek Results", "No results found from DeepSeek.")
                        return
                    
                    # Create and show the selection dialog
                    selection_dialog = DeepSeekSelectionDialog(dialog, parsed_codes, country)
                    dialog.wait_window(selection_dialog)  # Wait for user to close the dialog
                    
                    # Get the selected codes after dialog is closed
                    selected_codes = getattr(selection_dialog, 'selected_codes', [])
                    
                    if selected_codes:
                        saved_count = 0
                        for code in selected_codes:
                            if GUI_db.save_hs_code(code['hs_code'], code['description'], country, source='DeepSeek'):
                                saved_count += 1
                        messagebox.showinfo("Saved", f"Saved {saved_count} HS codes to the database.")
                        dialog.destroy()
                        self.populate_table()
                        
                        # Refresh other pages that depend on HS codes
                        self.refresh_other_pages()
                    else:
                        progress_label.configure(text="")
                        messagebox.showinfo("DeepSeek Results", "No HS codes were saved.")
                dialog.after(0, on_done)
            threading.Thread(target=run_deepseek, daemon=True).start()
        
        ctk.CTkButton(button_frame, text="Search with DeepSeek", fg_color="#4DA6FF", text_color="#121A26", font=("Poppins", 14, "bold"), command=on_deepseek_search).pack(side="left", padx=(0, 8))
        ctk.CTkButton(button_frame, text="Save", fg_color="#0078D4", text_color="#FFFFFF", font=("Poppins", 14, "bold"), command=on_manual_save).pack(side="left", padx=(0, 8))
        ctk.CTkButton(button_frame, text="Cancel", fg_color="#B0BEC5", text_color="#FFFFFF", font=("Poppins", 14), command=dialog.destroy).pack(side="right")

    def edit_selected(self):
        selected = self.table.selection()
        if not selected:
            messagebox.showwarning("Edit HS Code", "Please select a row to edit.")
            return
        hs_id = int(selected[0])
        entry = None
        for row in GUI_db.get_all_hs_codes():
            if row['id'] == hs_id:
                entry = row
                break
        if not entry:
            messagebox.showerror("Not Found", "Selected HS code not found.")
            return
        dialog = ctk.CTkToplevel(self)
        dialog.title("Edit HS Code")
        dialog.geometry("400x260")
        dialog.grab_set()
        ctk.CTkLabel(dialog, text="HS Code", font=("Poppins", 15)).pack(pady=(24, 4))
        hs_var = tk.StringVar(value=entry['hs_code'])
        hs_entry = ctk.CTkEntry(dialog, textvariable=hs_var, font=("Poppins", 15))
        hs_entry.pack(pady=4)
        ctk.CTkLabel(dialog, text="Description", font=("Poppins", 15)).pack(pady=(12, 4))
        desc_var = tk.StringVar(value=entry['description'])
        desc_entry = ctk.CTkEntry(dialog, textvariable=desc_var, font=("Poppins", 15))
        desc_entry.pack(pady=4)
        ctk.CTkLabel(dialog, text="Country", font=("Poppins", 15)).pack(pady=(12, 4))
        country_var = tk.StringVar(value=entry['country'])
        country_combo = ctk.CTkComboBox(dialog, variable=country_var, values=GUI_db.get_all_available_countries(), font=("Poppins", 15))
        country_combo.pack(pady=4)
        def on_save():
            hs = hs_var.get().strip()
            desc = desc_var.get().strip()
            country = country_var.get().strip()
            if not hs or not desc or not country:
                messagebox.showwarning("Missing Data", "Please fill in all fields.")
                return
            if GUI_db.update_hs_code(hs_id, hs, desc, country):
                dialog.destroy()
                self.populate_table()
                # Refresh other pages that depend on HS codes
                self.refresh_other_pages()
            else:
                messagebox.showerror("Error", "Failed to update HS code.")
        ctk.CTkButton(dialog, text="Save", fg_color="#0078D4", text_color="#FFFFFF", font=("Poppins", 15, "bold"), command=on_save).pack(pady=(18, 8))

    def delete_selected(self):
        selected = self.table.selection()
        if not selected:
            messagebox.showwarning("Delete HS Code", "Please select a row to delete.")
            return
        hs_id = int(selected[0])
        if messagebox.askyesno("Delete HS Code", "Are you sure you want to delete this HS code?"):
            if GUI_db.delete_hs_code(hs_id):
                self.populate_table()
                # Refresh other pages that depend on HS codes
                self.refresh_other_pages()
            else:
                messagebox.showerror("Error", "Failed to delete HS code.")

    def refresh_other_pages(self):
        """Refresh other pages that depend on HS codes"""
        try:
            # Invalidate cache
            cache.invalidate("buyer_search_countries")
            cache.invalidate("hs_codes")
            cache.invalidate("apollo_countries")
            cache.invalidate("apollo_companies")
            
            # Find the main app instance to access other pages
            main_app = self.winfo_toplevel()
            if hasattr(main_app, 'pages'):
                # Refresh BuyerSearchPage country list
                buyer_search_page = main_app.pages[1]  # Index 1 is BuyerSearchPage
                if hasattr(buyer_search_page, 'refresh_buyer_search_data'):
                    buyer_search_page.refresh_buyer_search_data()
        except Exception as e:
            print(f"Error refreshing other pages: {e}")



class DeepSeekSelectionDialog(ctk.CTkToplevel):
    def __init__(self, parent, parsed_codes, country):
        super().__init__(parent)
        self.title("Save DeepSeek HS Codes")
        self.geometry("650x600")  # Fixed height
        self.grab_set()
        
        self.parsed_codes = parsed_codes
        self.country = country
        self.selected_indices = []
        
        ctk.CTkLabel(self, text=f"Select which HS codes to save for {country}:", font=("Poppins", 18, "bold"), text_color="#2E3A59").pack(pady=(24, 16))
        
        # Select All checkbox
        self.select_all_var = tk.BooleanVar(value=True)
        select_all_cb = ctk.CTkCheckBox(self, text="Select All", variable=self.select_all_var, font=("Poppins", 15), command=self.toggle_all)
        select_all_cb.pack(anchor="w", padx=24, pady=(0, 16))
        
        # Scrollable frame for checkboxes
        scroll_frame = ctk.CTkScrollableFrame(self, height=300)
        scroll_frame.pack(fill="both", expand=True, padx=24, pady=(0, 16))
        
        self.checkboxes = []
        for i, code in enumerate(parsed_codes):
            cb = ctk.CTkCheckBox(
                scroll_frame, 
                text=f"HS Code: {code['hs_code']}\nDescription: {code['description']}", 
                variable=tk.BooleanVar(value=True),
                font=("Poppins", 14)
            )
            cb.pack(anchor="w", pady=4)
            self.checkboxes.append(cb)
        
        # Buttons with improved styling
        button_frame = ctk.CTkFrame(self, fg_color="#F5F7FA", height=80)
        button_frame.pack(fill="x", padx=24, pady=(16, 24))
        button_frame.pack_propagate(False)
        
        # Inner button container for better spacing
        inner_button_frame = ctk.CTkFrame(button_frame, fg_color="transparent")
        inner_button_frame.pack(fill="both", expand=True, padx=16, pady=16)
        
        # Left side - Cancel button
        ctk.CTkButton(
            inner_button_frame, 
            text="Cancel", 
            fg_color="#B0BEC5", 
            hover_color="#90A4AE",
            text_color="#FFFFFF", 
            font=("Poppins", 14, "bold"),
            width=120,
            height=36,
            command=self.reject
        ).pack(side="left")
        
        # Right side - Save button
        ctk.CTkButton(
            inner_button_frame, 
            text="Save Selected", 
            fg_color="#0078D4", 
            hover_color="#005A9E",
            text_color="#FFFFFF", 
            font=("Poppins", 14, "bold"),
            width=140,
            height=36,
            command=self.accept
        ).pack(side="right")
    
    def toggle_all(self):
        state = self.select_all_var.get()
        for cb in self.checkboxes:
            cb.select() if state else cb.deselect()
    
    def get_selected(self):
        selected = []
        for i, cb in enumerate(self.checkboxes):
            if cb.get():
                selected.append(self.parsed_codes[i])
        return selected
    
    def accept(self):
        self.selected_codes = self.get_selected()
        self.destroy()
    
    def reject(self):
        self.selected_codes = []
        self.destroy()


class ApolloPage(ctk.CTkFrame):
    def __init__(self, master):
        super().__init__(master, fg_color="#F5F7FA")
        self._build_ui()
        self.worker = None
        self.load_companies()
        self.load_countries()

    def _build_ui(self):
        # Title
        title_frame = ctk.CTkFrame(self, fg_color="#F5F7FA")
        title_frame.pack(fill="x", pady=(24, 16), padx=32)
        ctk.CTkLabel(title_frame, text="Potential Buyer Leads Search (Apollo)", font=("Poppins", 24, "bold"), text_color="#2E3A59").pack()
        
        # Search Potential Buyers' Companies button
        search_buyers_btn = ctk.CTkButton(title_frame, text="Search Potential Buyers' Companies", fg_color="#0078D4", hover_color="#005A9E", text_color="#FFFFFF", font=("Poppins", 16, "bold"), corner_radius=8, command=self.open_search_dialog)
        search_buyers_btn.pack(anchor="w", pady=(16, 0))
        
        # Search Parameters Card
        search_card = ctk.CTkFrame(self, fg_color="#FFFFFF", corner_radius=16)
        search_card.pack(fill="x", padx=32, pady=(0, 16))
        
        search_content = ctk.CTkFrame(search_card, fg_color="transparent")
        search_content.pack(fill="x", padx=20, pady=18)
        
        ctk.CTkLabel(search_content, text="Search Parameters", font=("Poppins", 18, "bold"), text_color="#0078D4").pack(anchor="w", pady=(0, 12))
        
        # Parameters row (now as a grid for better layout)
        param_frame = ctk.CTkFrame(search_content, fg_color="transparent")
        param_frame.pack(fill="x", pady=8)
        param_frame.grid_columnconfigure(0, weight=1)
        param_frame.grid_columnconfigure(1, weight=1)

        # Country selection (row 0)
        ctk.CTkLabel(param_frame, text="Country:", font=("Poppins", 15), text_color="#6B7C93").grid(row=0, column=0, sticky="w")
        self.country_var = tk.StringVar(value="Select Country")
        country_frame = ctk.CTkFrame(param_frame, fg_color="transparent")
        country_frame.grid(row=0, column=1, sticky="ew", padx=(8, 24))
        self.country_display = ctk.CTkEntry(country_frame, textvariable=self.country_var, state="readonly", width=200, font=("Poppins", 15))
        self.country_display.pack(side="left")
        select_country_btn = ctk.CTkButton(country_frame, text="Select", fg_color="#4CAF50", hover_color="#388E3C", text_color="#FFFFFF", font=("Poppins", 12, "bold"), width=60, height=32, command=self.open_country_selector)
        select_country_btn.pack(side="left", padx=(8, 0))

        # Company selection (row 1)
        ctk.CTkLabel(param_frame, text="Company:", font=("Poppins", 15), text_color="#6B7C93").grid(row=1, column=0, sticky="w", pady=(8,0))
        self.company_var = tk.StringVar(value="Type or select a company...")
        company_frame = ctk.CTkFrame(param_frame, fg_color="transparent")
        company_frame.grid(row=1, column=1, sticky="ew", padx=(8, 24), pady=(8,0))
        self.company_display = ctk.CTkEntry(company_frame, textvariable=self.company_var, state="readonly", width=280, font=("Poppins", 15))
        self.company_display.pack(side="left")
        self.select_btn = ctk.CTkButton(company_frame, text="Select", fg_color="#4CAF50", hover_color="#388E3C", text_color="#FFFFFF", font=("Poppins", 12, "bold"), width=60, height=32, command=self.open_company_selector)
        self.select_btn.pack(side="left", padx=(8, 0))
        # Initially disable company entry and button
        self.company_display.configure(state="disabled")
        self.select_btn.configure(state="disabled")

        # Update enable/disable state when country changes
        def on_country_change(*args):
            country = self.country_var.get()
            if country and country != "Select Country":
                self.company_display.configure(state="readonly")
                self.select_btn.configure(state="normal")
            else:
                self.company_display.configure(state="disabled")
                self.select_btn.configure(state="disabled")
        self.country_var.trace_add('write', on_country_change)

        # Search button (row 2)
        self.search_btn = ctk.CTkButton(param_frame, text="Search Decision Makers", fg_color="#0078D4", hover_color="#005A9E", text_color="#FFFFFF", font=("Poppins", 15, "bold"), corner_radius=8, command=self.do_search)
        self.search_btn.grid(row=2, column=0, columnspan=2, sticky="e", pady=(12,0))
        
        # Results Card
        results_card = ctk.CTkFrame(self, fg_color="#FFFFFF", corner_radius=16)
        results_card.pack(fill="both", expand=True, padx=32, pady=(0, 24))
        
        results_content = ctk.CTkFrame(results_card, fg_color="transparent")
        results_content.pack(fill="both", expand=True, padx=20, pady=18)
        
        # Info and action buttons row
        info_action_frame = ctk.CTkFrame(results_content, fg_color="transparent")
        info_action_frame.pack(fill="x", pady=(0, 12))
        
        self.results_info_label = ctk.CTkLabel(info_action_frame, text="", font=("Poppins", 15), text_color="#0078D4", fg_color="#EAF3FF", corner_radius=8)
        self.results_info_label.pack(side="left", padx=(0, 16))
        
        # Action buttons
        duplicate_btn = ctk.CTkButton(info_action_frame, text="Check for Duplicate", fg_color="#E0E4EA", hover_color="#B0BEC5", text_color="#2E3A59", font=("Poppins", 13, "bold"), corner_radius=6, width=180, height=40, command=self.check_duplicates)
        duplicate_btn.pack(side="right", padx=(0, 24))  # Increased spacing to 24 pixels
        
        export_btn = ctk.CTkButton(info_action_frame, text="Export Results", fg_color="#0078D4", hover_color="#005A9E", text_color="#FFFFFF", font=("Poppins", 15, "bold"), corner_radius=8, width=180, height=40, command=self.export_results)
        export_btn.pack(side="right", padx=(0, 8))  # Add some right padding to the export button
        
        # Table
        table_frame = ctk.CTkFrame(results_content, fg_color="#FFFFFF", corner_radius=12)
        table_frame.pack(fill="both", expand=True, pady=(0, 8))
        
        style = ttk.Style()
        style.configure("Apollo.Treeview", font=("Poppins", 13), rowheight=32, background="#FFFFFF", fieldbackground="#FFFFFF", foreground="#2E3A59")
        style.configure("Apollo.Treeview.Heading", font=("Poppins", 14, "bold"), background="#F5F7FA", foreground="#2E3A59")
        style.map("Apollo.Treeview", background=[("selected", "#1E3A8A")], foreground=[("selected", "#FFFFFF")])
        
        self.table = ttk.Treeview(table_frame, columns=("no", "name", "title", "email", "linkedin", "company"), show="headings", style="Apollo.Treeview")
        self.table.heading("no", text="No.")
        self.table.heading("name", text="Name")
        self.table.heading("title", text="Title")
        self.table.heading("email", text="Email")
        self.table.heading("linkedin", text="LinkedIn")
        self.table.heading("company", text="Company")
        self.table.column("no", width=60, anchor="center")
        self.table.column("name", width=150, anchor="w")
        self.table.column("title", width=200, anchor="w")
        self.table.column("email", width=200, anchor="w")
        self.table.column("linkedin", width=150, anchor="w")
        self.table.column("company", width=200, anchor="w")
        self.table.pack(fill="both", expand=True, padx=8, pady=8)
        
        # Progress bar
        self.progress_bar = ctk.CTkProgressBar(results_content, orientation="horizontal", mode="indeterminate", width=320)
        self.progress_bar.pack(pady=8)
        self.progress_bar.set(0)
        self.progress_bar.pack_forget()  # Hide initially
        
        # Initialize data
        self.load_companies()
        self.load_countries()
        self.update_company_completer()

    def update_company_completer(self, *args):
        """Update the company completer based on selected country"""
        try:
            from GUI_db import get_all_companies
            selected_country = self.country_var.get().strip()
            companies = get_all_companies()
            company_names = ["Type or select a company..."]
            for company in companies:
                if selected_country == "Select Country" or not selected_country or company.get('country', '').strip() == selected_country:
                    name = company.get('company_name', '')
                    if name:
                        company_names.append(name)
            self.filtered_companies = company_names
            self.company_var.set("Type or select a company...")
        except Exception as e:
            print(f"Error updating company completer: {e}")

    def load_countries(self):
        """Load available countries from database that have companies"""
        try:
            companies = GUI_db.get_all_companies()
            countries_with_companies = set()
            for company in companies:
                if company.get('country'):
                    countries_with_companies.add(company['country'])
            countries = ["Select Country"] + sorted(list(countries_with_companies))
            self.country_list = countries
        except Exception as e:
            print(f"Error loading Apollo countries: {e}")
            self.country_list = ["Select Country"]

    def open_country_selector(self):
        """Open a searchable country selection dialog"""
        dialog = CountrySelectorDialog(self, self.country_list)
        self.wait_window(dialog)  # Wait for dialog to close
        selected_country = dialog.get_selected()
        if selected_country and selected_country != "Select Country":
            self.country_var.set(selected_country)
            # Update company completer when country changes
            self.update_company_completer()

    def open_company_selector(self):
        """Open a searchable company selection dialog"""
        dialog = CompanySelectorDialog(self, self.filtered_companies if hasattr(self, 'filtered_companies') else ["Type or select a company..."])
        self.wait_window(dialog)  # Wait for dialog to close
        selected_company = dialog.get_selected()
        if selected_company and selected_company != "Type or select a company...":
            self.company_var.set(selected_company)
    

    def load_companies(self):
        """Load companies from database"""
        self.update_company_completer()

    def refresh_apollo_data(self):
        """Refresh Apollo page data after new companies are added"""
        self.load_countries()
        self.load_companies()

    def do_search(self):
        """Perform Apollo search for decision makers"""
        company_name = self.company_var.get().strip()
        country = self.country_var.get().strip()
        
        if not company_name or company_name == "Type or select a company...":
            messagebox.showwarning("Missing Company", "Please select or enter a company name.")
            return
            
        if country == "Select Country" or not country:
            messagebox.showwarning("Missing Country", "Please select a country.")
            return
        
        # Start worker thread
        self.search_btn.configure(state="disabled")
        self.progress_bar.pack(pady=8)
        self.progress_bar.start()
        
        def run_apollo_search():
            try:
                from apollo import find_decision_makers_apollo
                from GUI_db import insert_contact, get_all_companies, insert_company
                
                results = find_decision_makers_apollo(company_name, country)
                
                # Save results to database if any found
                if results:
                    # Find the company ID in the database
                    companies = get_all_companies()
                    company_id = None
                    for company in companies:
                        if company.get('company_name', '').lower() == company_name.lower():
                            company_id = company.get('id')
                            break
                    
                    # If company not found, create it
                    if not company_id:
                        company_id, _ = insert_company(company_name, country, '', '', 0)
                    
                    # Save each contact
                    saved_count = 0
                    for contact in results:
                        try:
                            insert_contact(
                                company_id,
                                company_name,
                                contact.get('name', ''),
                                contact.get('title', ''),
                                contact.get('email', ''),
                                contact.get('linkedin', '')
                            )
                            saved_count += 1
                        except Exception as e:
                            print(f"Error saving contact: {e}")
                
                # Add company name to each result for display
                for contact in results:
                    contact['company_name'] = company_name
                
                def on_complete():
                    self.search_btn.configure(state="normal")
                    self.progress_bar.stop()
                    self.progress_bar.pack_forget()
                    
                    if not results:
                        self.results_info_label.configure(text="")
                        messagebox.showinfo("No Results", "No decision makers found for this company.")
                        return
                    
                    self.populate_table(results)
                    from utils.display import truncate_company_name
                    truncated_company = truncate_company_name(company_name)
                    self.results_info_label.configure(text=f"Number of {len(results)} buyer leads found for {truncated_company}")
                    
                    # Auto-refresh country and company lists after successful search
                    self.load_countries()
                    self.load_companies()
                    
                    messagebox.showinfo("Search Complete", f"Found and saved {len(results)} decision makers to the database.")
                
                self.after(0, on_complete)
                
            except Exception as e:
                def on_error():
                    self.search_btn.configure(state="normal")
                    self.progress_bar.stop()
                    self.progress_bar.pack_forget()
                    messagebox.showerror("Search Error", f"Error during search: {str(e)}")
                self.after(0, on_error)
        
        threading.Thread(target=run_apollo_search, daemon=True).start()

    def populate_table(self, data):
        """Populate the results table"""
        for row in self.table.get_children():
            self.table.delete(row)
        
        for i, contact in enumerate(data):
            from utils.display import truncate_company_name
            company_name = contact.get('company_name', '')
            truncated_company = truncate_company_name(company_name, max_length=25)
            self.table.insert("", "end", values=(
                str(i + 1),
                contact.get('name', ''),
                contact.get('title', ''),
                contact.get('email', ''),
                contact.get('linkedin', ''),
                truncated_company
            ))

    def export_results(self):
        """Export results to CSV"""
        if not self.table.get_children():
            messagebox.showwarning("No Data", "No results to export.")
            return
            
        try:
            # Generate default filename based on selected company
            selected_company = self.company_var.get()
            if selected_company and selected_company != "Type or select a company...":
                # Clean company name for filename (remove special characters)
                clean_company = "".join(c for c in selected_company if c.isalnum() or c in (' ', '-', '_')).rstrip()
                default_filename = f"{clean_company}_buyer_list.csv"
            else:
                default_filename = "apollo_buyer_list.csv"
            
            filename = filedialog.asksaveasfilename(
                defaultextension=".csv",
                filetypes=[("CSV files", "*.csv"), ("All files", "*.*")],
                title="Export Results",
                initialfile=default_filename
            )
            
            if filename:
                with open(filename, 'w', newline='', encoding='utf-8') as csvfile:
                    writer = csv.writer(csvfile)
                    # Write headers
                    headers = ["No.", "Name", "Title", "Email", "LinkedIn", "Company"]
                    writer.writerow(headers)
                    
                    # Write data
                    for item in self.table.get_children():
                        values = self.table.item(item)['values']
                        writer.writerow(values)
                
                messagebox.showinfo("Export Complete", f"Results exported to {filename}")
                
        except Exception as e:
            messagebox.showerror("Export Error", f"Error exporting results: {e}")

    def open_search_dialog(self):
        """Open the Apollo search dialog"""
        dialog = ApolloSearchDialog(self)
        # The refresh will be handled by the dialog's on_complete callback

    def check_duplicates(self):
        """Check the whole database for duplicate companies"""
        try:
            from GUI_db import get_all_companies
            companies = get_all_companies()
            seen = {}
            duplicates = []
            for company in companies:
                key = (company.get('company_name', '').strip().lower(), company.get('country', '').strip().lower())
                if key in seen:
                    duplicates.append(company)
                else:
                    seen[key] = company
            
            if duplicates:
                msg = "Duplicate companies found (same name and country):\n\n"
                for dup in duplicates:
                    msg += f"- {dup.get('company_name', '')} ({dup.get('country', '')})\n"
                messagebox.showwarning("Duplicate Companies", msg)
            else:
                messagebox.showinfo("Duplicate Companies", "No duplicate companies found in the database.")
        except Exception as e:
            messagebox.showerror("Error", f"Error checking duplicates: {e}")


class ApolloSearchDialog(ctk.CTkToplevel):
    def __init__(self, parent):
        super().__init__(parent)
        self.title("Search Potential Buyers' Companies")
        self.geometry("500x400")
        self.grab_set()
        
        # Initialize Apollo database
        try:
            from GUI_db import init_apollo_db
            init_apollo_db()
        except Exception as e:
            print(f"Error initializing Apollo database: {e}")
        
        self._build_ui()

    def _build_ui(self):
        ctk.CTkLabel(self, text="Search Potential Buyers' Companies", font=("Poppins", 18, "bold"), text_color="#2E3A59").pack(pady=(24, 16))
        
        # Search parameters
        params_frame = ctk.CTkFrame(self, fg_color="#F5F7FA")
        params_frame.pack(fill="x", padx=24, pady=8)
        
        ctk.CTkLabel(params_frame, text="Search Parameters", font=("Poppins", 16, "bold"), text_color="#0078D4").pack(anchor="w", pady=(16, 12))
        
        # Country selection
        country_frame = ctk.CTkFrame(params_frame, fg_color="transparent")
        country_frame.pack(fill="x", padx=16, pady=4)
        ctk.CTkLabel(country_frame, text="Country:", font=("Poppins", 14)).pack(side="left")
        self.country_var = tk.StringVar(value="Select Country")
        country_display_frame = ctk.CTkFrame(country_frame, fg_color="transparent")
        country_display_frame.pack(side="right", fill="x", expand=True, padx=(8, 0))
        self.country_display = ctk.CTkEntry(country_display_frame, textvariable=self.country_var, state="readonly", font=("Poppins", 14))
        self.country_display.pack(side="left", fill="x", expand=True)
        select_country_btn = ctk.CTkButton(country_display_frame, text="Select", fg_color="#4CAF50", hover_color="#388E3C", text_color="#FFFFFF", font=("Poppins", 12, "bold"), width=60, height=32, command=self.open_country_selector)
        select_country_btn.pack(side="right", padx=(8, 0))
        
        # Search depth
        depth_frame = ctk.CTkFrame(params_frame, fg_color="transparent")
        depth_frame.pack(fill="x", padx=16, pady=4)
        ctk.CTkLabel(depth_frame, text="Search Depth:", font=("Poppins", 14)).pack(side="left")
        self.depth_var = tk.StringVar(value="Medium (1,000 companies)")
        depth_combo = ctk.CTkComboBox(depth_frame, variable=self.depth_var, values=[
            "Small (500 companies)", "Medium (1,000 companies)", "Large (2,500 companies)", 
            "Very Large (5,000 companies)", "Maximum (10,000 companies)"
        ], font=("Poppins", 14))
        depth_combo.pack(side="right", fill="x", expand=True, padx=(8, 0))
        
        # Progress bar
        self.progress_bar = ctk.CTkProgressBar(self, orientation="horizontal", mode="indeterminate", width=320)
        self.progress_bar.pack(pady=16)
        self.progress_bar.set(0)
        self.progress_bar.pack_forget()  # Hide initially
        
        # Buttons
        button_frame = ctk.CTkFrame(self, fg_color="#F5F7FA")
        button_frame.pack(fill="x", padx=24, pady=(16, 24))
        
        ctk.CTkButton(button_frame, text="Search", fg_color="#0078D4", text_color="#FFFFFF", font=("Poppins", 14, "bold"), command=self.start_search).pack(side="left", padx=(0, 8))
        ctk.CTkButton(button_frame, text="Cancel", fg_color="#B0BEC5", text_color="#FFFFFF", font=("Poppins", 14), command=self.destroy).pack(side="right")

    def open_country_selector(self):
        """Open a searchable country selection dialog"""
        country_list = ["Select Country"] + self._get_all_countries()
        dialog = CountrySelectorDialog(self, country_list)
        self.wait_window(dialog)  # Wait for dialog to close
        selected_country = dialog.get_selected()
        if selected_country and selected_country != "Select Country":
            self.country_var.set(selected_country)

    def _get_all_countries(self):
        """Get all available countries"""
        return [
            # Global countries
            'United States', 'Germany', 'United Kingdom', 'France', 'Italy', 'Spain', 'Canada', 'Australia', 'Brazil', 'Mexico',
            'Russia', 'Turkey', 'Netherlands', 'Switzerland', 'Sweden', 'Norway', 'Denmark', 'Finland', 'Poland', 'Austria',
            'Belgium', 'South Africa', 'Egypt', 'Saudi Arabia', 'United Arab Emirates', 'Argentina', 'Chile', 'New Zealand',
            'Ireland', 'Portugal', 'Greece', 'Czech Republic', 'Hungary', 'Romania', 'Israel', 'Ukraine',
            # Asia countries
            'Malaysia', 'Indonesia', 'Thailand', 'Vietnam', 'Singapore', 'Philippines', 'China', 'India', 'Japan', 'South Korea',
            'Hong Kong', 'Taiwan', 'Bangladesh', 'Pakistan', 'Sri Lanka', 'Myanmar', 'Cambodia', 'Laos', 'Nepal', 'Mongolia',
            'Brunei', 'Timor-Leste', 'Maldives', 'Bhutan'
        ]

    def start_search(self):
        """Start the Apollo company search"""
        country = self.country_var.get().strip()
        if country == "Select Country" or not country:
            messagebox.showwarning("Missing Country", "Please select a country.")
            return
        depth_text = self.depth_var.get()
        
        # Parse search depth
        depth_map = {
            "Small (500 companies)": 5,
            "Medium (1,000 companies)": 10,
            "Large (2,500 companies)": 25,
            "Very Large (5,000 companies)": 50,
            "Maximum (10,000 companies)": 100
        }
        max_pages = depth_map.get(depth_text, 10)
        
        self.progress_bar.pack(pady=16)
        self.progress_bar.start()
        
        def run_search():
            try:
                from apollo import APOLLO_API_KEY
                import requests
                from GUI_db import insert_company, count_companies
                
                if not APOLLO_API_KEY:
                    raise Exception("APOLLO_API_KEY environment variable not set!")
                
                headers = {
                    "Cache-Control": "no-cache", 
                    "Content-Type": "application/json", 
                    "x-api-key": APOLLO_API_KEY
                }
                
                initial_count = count_companies()
                total_saved = 0
                
                for page in range(1, max_pages + 1):
                    body = {
                        "page": page,
                        "per_page": 100,
                        "industry_tags": [
                            "Pharmaceuticals", "Medical Devices", "Healthcare",
                            "Manufacturing", "Medical Supplies"
                        ],
                        "q_organization_keyword_tags": [
                            "latex gloves", "nitrile gloves", "medical gloves",
                             "exam gloves", "biohazard protection", "disposable gloves"
                        ]
                    }
                    
                    if country:
                        body["organization_locations"] = [country]
                    
                    resp = requests.post(
                        "https://api.apollo.io/v1/mixed_companies/search",
                        headers=headers,
                        json=body
                    )
                    
                    if resp.status_code != 200:
                        raise Exception(f"Apollo API error (status {resp.status_code}): {resp.text}")
                    
                    data = resp.json()
                    companies = data.get("organizations", [])
                    
                    if not companies:
                        break
                    
                    for comp in companies:
                        company_name = comp.get("name", "")
                        country_val = comp.get("location_country") or comp.get("country") or country
                        domain = comp.get("primary_domain")
                        industry = comp.get("industry", "")
                        employee_count = comp.get("estimated_num_employees")
                        
                        if not company_name or not domain:
                            continue
                        
                        cid, is_new = insert_company(company_name, country_val, domain, industry, employee_count)
                        if is_new:
                            total_saved += 1
                
                def on_complete():
                    self.progress_bar.stop()
                    self.progress_bar.pack_forget()
                    messagebox.showinfo("Search Complete", f"Found and saved {total_saved} new companies to the database.")
                    
                    # Refresh the parent Apollo page after successful search
                    try:
                        parent = self.master
                        while parent and not hasattr(parent, 'refresh_apollo_data'):
                            parent = parent.master
                        if parent and hasattr(parent, 'refresh_apollo_data'):
                            parent.refresh_apollo_data()
                    except Exception:
                        pass  # Ignore if refresh fails
                    
                    self.destroy()
                
                self.after(0, on_complete)
                
            except Exception as e:
                def on_error():
                    self.progress_bar.stop()
                    self.progress_bar.pack_forget()
                    messagebox.showerror("Search Error", f"Error during search: {str(e)}")
                self.after(0, on_error)
        
        threading.Thread(target=run_search, daemon=True).start()


class CountrySelectorDialog(ctk.CTkToplevel):
    def __init__(self, parent, country_list):
        super().__init__(parent)
        self.title("Select Country")
        self.geometry("500x600")  # Increased height
        self.grab_set()
        self.country_list = country_list
        self.selected_country = None
        self.current_page = 0
        self.items_per_page = 100
        ctk.CTkLabel(self, text="Select a country:", font=("Poppins", 16, "bold"), text_color="#2E3A59").pack(pady=(24, 12))
        search_frame = ctk.CTkFrame(self, fg_color="#F5F7FA")
        search_frame.pack(fill="x", padx=24, pady=(0, 12))
        self.search_var = tk.StringVar()
        self.search_var.trace("w", self.on_search)
        search_entry = ctk.CTkEntry(search_frame, textvariable=self.search_var, placeholder_text="Search countries...", font=("Poppins", 14))
        search_entry.pack(fill="x", padx=16, pady=16)
        self.scroll_frame = ctk.CTkScrollableFrame(self, height=250)
        self.scroll_frame.pack(fill="both", expand=True, padx=24, pady=(0, 12))
        self.country_buttons = []
        self.filtered_countries = self.country_list[:]
        # Pagination controls (must be created before populate_countries)
        self.pagination_frame = ctk.CTkFrame(self, fg_color="#F5F7FA")
        self.pagination_frame.pack(fill="x", padx=24, pady=(0, 0))
        self.prev_btn = ctk.CTkButton(self.pagination_frame, text="Previous", command=self.prev_page, width=80)
        self.next_btn = ctk.CTkButton(self.pagination_frame, text="Next", command=self.next_page, width=80)
        self.page_label = ctk.CTkLabel(self.pagination_frame, text="")
        self.prev_btn.pack(side="left")
        self.page_label.pack(side="left", padx=8)
        self.next_btn.pack(side="left")
        button_frame = ctk.CTkFrame(self, fg_color="#F5F7FA")
        button_frame.pack(fill="x", padx=24, pady=(0, 24))
        ctk.CTkButton(button_frame, text="Cancel", fg_color="#B0BEC5", text_color="#FFFFFF", font=("Poppins", 14), command=self.destroy).pack(side="right")
        # Now safe to call populate_countries
        self.populate_countries()
        self.update_pagination_controls()

    def populate_countries(self):
        for btn in self.country_buttons:
            btn.destroy()
        self.country_buttons.clear()
        start = self.current_page * self.items_per_page
        end = start + self.items_per_page
        shown_countries = self.filtered_countries[start:end]
        for country in shown_countries:
            if country != "Select Country":
                from utils.display import truncate_company_name
                truncated_country = truncate_company_name(country, max_length=35)
                btn = ctk.CTkButton(
                    self.scroll_frame,
                    text=truncated_country, 
                    fg_color="#FFFFFF", 
                    hover_color="#EAF3FF",
                    text_color="#2E3A59", 
                    font=("Poppins", 13),
                    anchor="w",
                    command=lambda c=country: self.select_country(c)
                )
                btn.pack(fill="x", padx=8, pady=2)
                self.country_buttons.append(btn)
        self.update_pagination_controls()

    def update_pagination_controls(self):
        total_pages = max(1, (len(self.filtered_countries) - 1) // self.items_per_page + 1)
        self.page_label.configure(text=f"Page {self.current_page + 1} of {total_pages}")
        self.prev_btn.configure(state="normal" if self.current_page > 0 else "disabled")
        self.next_btn.configure(state="normal" if (self.current_page + 1) * self.items_per_page < len(self.filtered_countries) else "disabled")

    def next_page(self):
        if (self.current_page + 1) * self.items_per_page < len(self.filtered_countries):
            self.current_page += 1
            self.populate_countries()

    def prev_page(self):
        if self.current_page > 0:
            self.current_page -= 1
            self.populate_countries()

    def on_search(self, *args):
        search_term = self.search_var.get().lower()
        self.filtered_countries = [country for country in self.country_list if search_term in country.lower()]
        self.current_page = 0
        self.populate_countries()

    def select_country(self, country):
        self.selected_country = country
        self.destroy()

    def get_selected(self):
        return self.selected_country

class CompanySelectorDialog(ctk.CTkToplevel):
    def __init__(self, parent, company_list):
        super().__init__(parent)
        self.title("Select Company")
        self.geometry("500x600")  # Increased height
        self.grab_set()
        self.company_list = company_list
        self.selected_company = None
        self.current_page = 0
        self.items_per_page = 100
        ctk.CTkLabel(self, text="Select a company:", font=("Poppins", 16, "bold"), text_color="#2E3A59").pack(pady=(24, 12))
        search_frame = ctk.CTkFrame(self, fg_color="#F5F7FA")
        search_frame.pack(fill="x", padx=24, pady=(0, 12))
        self.search_var = tk.StringVar()
        self.search_var.trace("w", self.on_search)
        search_entry = ctk.CTkEntry(search_frame, textvariable=self.search_var, placeholder_text="Search companies...", font=("Poppins", 14))
        search_entry.pack(fill="x", padx=16, pady=16)
        self.scroll_frame = ctk.CTkScrollableFrame(self, height=250)
        self.scroll_frame.pack(fill="both", expand=True, padx=24, pady=(0, 12))
        self.company_buttons = []
        self.filtered_companies = self.company_list[:]
        # Pagination controls (must be created before populate_companies)
        self.pagination_frame = ctk.CTkFrame(self, fg_color="#F5F7FA")
        self.pagination_frame.pack(fill="x", padx=24, pady=(0, 0))
        self.prev_btn = ctk.CTkButton(self.pagination_frame, text="Previous", command=self.prev_page, width=80)
        self.next_btn = ctk.CTkButton(self.pagination_frame, text="Next", command=self.next_page, width=80)
        self.page_label = ctk.CTkLabel(self.pagination_frame, text="")
        self.prev_btn.pack(side="left")
        self.page_label.pack(side="left", padx=8)
        self.next_btn.pack(side="left")
        button_frame = ctk.CTkFrame(self, fg_color="#F5F7FA")
        button_frame.pack(fill="x", padx=24, pady=(0, 24))
        ctk.CTkButton(button_frame, text="Cancel", fg_color="#B0BEC5", text_color="#FFFFFF", font=("Poppins", 14), command=self.destroy).pack(side="right")
        # Now safe to call populate_companies
        self.populate_companies()
        self.update_pagination_controls()

    def populate_companies(self):
        for btn in self.company_buttons:
            btn.destroy()
        self.company_buttons.clear()
        start = self.current_page * self.items_per_page
        end = start + self.items_per_page
        shown_companies = self.filtered_companies[start:end]
        for company in shown_companies:
            if company != "Type or select a company...":
                from utils.display import truncate_company_name
                truncated_company = truncate_company_name(company, max_length=35)
                btn = ctk.CTkButton(
                    self.scroll_frame,  # Create button in scroll_frame, not self
                    text=truncated_company, 
                    fg_color="#FFFFFF", 
                    hover_color="#EAF3FF",
                    text_color="#2E3A59", 
                    font=("Poppins", 13),
                    anchor="w",
                    command=lambda c=company: self.select_company(c)
                )
                btn.pack(fill="x", padx=8, pady=2)
                self.company_buttons.append(btn)
        self.update_pagination_controls()

    def update_pagination_controls(self):
        total_pages = max(1, (len(self.filtered_companies) - 1) // self.items_per_page + 1)
        self.page_label.configure(text=f"Page {self.current_page + 1} of {total_pages}")
        self.prev_btn.configure(state="normal" if self.current_page > 0 else "disabled")
        self.next_btn.configure(state="normal" if (self.current_page + 1) * self.items_per_page < len(self.filtered_companies) else "disabled")

    def next_page(self):
        if (self.current_page + 1) * self.items_per_page < len(self.filtered_companies):
            self.current_page += 1
            self.populate_companies()

    def prev_page(self):
        if self.current_page > 0:
            self.current_page -= 1
            self.populate_companies()

    def on_search(self, *args):
        search_term = self.search_var.get().lower()
        self.filtered_companies = [company for company in self.company_list if search_term in company.lower()]
        self.current_page = 0
        self.populate_companies()

    def select_company(self, company):
        self.selected_company = company
        self.destroy()

    def get_selected(self):
        return self.selected_company





class HSCodeSelectorDialog(ctk.CTkToplevel):
    def __init__(self, parent, hs_code_list):
        super().__init__(parent)
        self.title("Select HS Code")
        self.geometry("600x400")
        self.grab_set()
        
        self.hs_code_list = hs_code_list
        self.selected_hs_code = None
        
        ctk.CTkLabel(self, text="Select an HS Code:", font=("Poppins", 16, "bold"), text_color="#2E3A59").pack(pady=(24, 12))
        
        # Search frame
        search_frame = ctk.CTkFrame(self, fg_color="#F5F7FA")
        search_frame.pack(fill="x", padx=24, pady=(0, 12))
        
        self.search_var = tk.StringVar()
        self.search_var.trace("w", self.on_search)
        search_entry = ctk.CTkEntry(search_frame, textvariable=self.search_var, placeholder_text="Search HS codes...", font=("Poppins", 14))
        search_entry.pack(fill="x", padx=16, pady=16)
        
        # Scrollable frame for HS codes
        self.scroll_frame = ctk.CTkScrollableFrame(self, height=250)
        self.scroll_frame.pack(fill="both", expand=True, padx=24, pady=(0, 12))
        
        # HS code buttons (will be populated)
        self.hs_code_buttons = []
        self.populate_hs_codes(self.hs_code_list)
        
        # Buttons
        button_frame = ctk.CTkFrame(self, fg_color="#F5F7FA")
        button_frame.pack(fill="x", padx=24, pady=(0, 24))
        
        ctk.CTkButton(button_frame, text="Cancel", fg_color="#B0BEC5", text_color="#FFFFFF", font=("Poppins", 14), command=self.destroy).pack(side="right")

    def populate_hs_codes(self, hs_codes):
        """Populate the scrollable frame with HS code buttons"""
        # Clear existing buttons
        for btn in self.hs_code_buttons:
            btn.destroy()
        self.hs_code_buttons.clear()
        
        # Add HS code buttons
        for hs_code in hs_codes:
            if hs_code != "Select HS Code":
                # Truncate long descriptions for display
                display_text = hs_code
                if len(display_text) > 60:
                    display_text = display_text[:57] + "..."
                
                btn = ctk.CTkButton(
                    self.scroll_frame,
                    text=display_text, 
                    fg_color="#FFFFFF", 
                    hover_color="#EAF3FF",
                    text_color="#2E3A59", 
                    font=("Poppins", 12),
                    anchor="w",
                    command=lambda hc=hs_code: self.select_hs_code(hc)
                )
                btn.pack(fill="x", padx=8, pady=2)
                self.hs_code_buttons.append(btn)

    def on_search(self, *args):
        """Filter HS codes based on search term"""
        search_term = self.search_var.get().lower()
        filtered_hs_codes = []
        
        for hs_code in self.hs_code_list:
            if search_term in hs_code.lower():
                filtered_hs_codes.append(hs_code)
        
        self.populate_hs_codes(filtered_hs_codes)

    def select_hs_code(self, hs_code):
        """Select an HS code and close dialog"""
        self.selected_hs_code = hs_code
        self.destroy()

    def get_selected(self):
        """Return the selected HS code"""
        return self.selected_hs_code


class KeywordSelectorDialog(ctk.CTkToplevel):
    def __init__(self, parent, keyword_list):
        super().__init__(parent)
        self.title("Select Keyword")
        self.geometry("500x400")
        self.grab_set()
        
        self.keyword_list = keyword_list
        self.selected_keyword = None
        
        ctk.CTkLabel(self, text="Select a keyword:", font=("Poppins", 16, "bold"), text_color="#2E3A59").pack(pady=(24, 12))
        
        # Search frame
        search_frame = ctk.CTkFrame(self, fg_color="#F5F7FA")
        search_frame.pack(fill="x", padx=24, pady=(0, 12))
        
        self.search_var = tk.StringVar()
        self.search_var.trace("w", self.on_search)
        search_entry = ctk.CTkEntry(search_frame, textvariable=self.search_var, placeholder_text="Search keywords...", font=("Poppins", 14))
        search_entry.pack(fill="x", padx=16, pady=16)
        
        # Scrollable frame for keywords
        self.scroll_frame = ctk.CTkScrollableFrame(self, height=250)
        self.scroll_frame.pack(fill="both", expand=True, padx=24, pady=(0, 12))
        
        # Keyword buttons (will be populated)
        self.keyword_buttons = []
        self.populate_keywords(self.keyword_list)
        
        # Buttons
        button_frame = ctk.CTkFrame(self, fg_color="#F5F7FA")
        button_frame.pack(fill="x", padx=24, pady=(0, 24))
        
        ctk.CTkButton(button_frame, text="Cancel", fg_color="#B0BEC5", text_color="#FFFFFF", font=("Poppins", 14), command=self.destroy).pack(side="right")

    def populate_keywords(self, keywords):
        """Populate the scrollable frame with keyword buttons"""
        # Clear existing buttons
        for btn in self.keyword_buttons:
            btn.destroy()
        self.keyword_buttons.clear()
        
        # Add keyword buttons
        for keyword in keywords:
            if keyword != "Select Keyword":
                btn = ctk.CTkButton(
                    self.scroll_frame,
                    text=keyword, 
                    fg_color="#FFFFFF", 
                    hover_color="#EAF3FF",
                    text_color="#2E3A59", 
                    font=("Poppins", 13),
                    anchor="w",
                    command=lambda k=keyword: self.select_keyword(k)
                )
                btn.pack(fill="x", padx=8, pady=2)
                self.keyword_buttons.append(btn)

    def on_search(self, *args):
        """Filter keywords based on search term"""
        search_term = self.search_var.get().lower()
        filtered_keywords = []
        
        for keyword in self.keyword_list:
            if search_term in keyword.lower():
                filtered_keywords.append(keyword)
        
        self.populate_keywords(filtered_keywords)

    def select_keyword(self, keyword):
        """Select a keyword and close dialog"""
        self.selected_keyword = keyword
        self.destroy()

    def get_selected(self):
        """Return the selected keyword"""
        return self.selected_keyword


class MainApp(ctk.CTk):
    def __init__(self):
        super().__init__()
        self.title("Glove Buyer App")
        # Launch in fullscreen more reliably
        self.after(10, self.maximize_window)
        self.configure(bg="#F5F7FA")
        # Sidebar with increased width for longer names
        self.sidebar = ctk.CTkFrame(self, fg_color="#F5F7FA", width=280, corner_radius=0)
        self.sidebar.pack(side="left", fill="y")
        self.sidebar.pack_propagate(False)
        # Branding/logo placeholder
        ctk.CTkLabel(self.sidebar, text="LOGO", font=("Poppins", 22, "bold"), text_color="#0078D4").pack(pady=(32, 24))
        # Nav buttons
        self.nav_btns = []
        for i, label in enumerate(NAV_LABELS):
            btn = ctk.CTkButton(
                self.sidebar, 
                text=label, 
                font=("Poppins", 15, "bold"), 
                fg_color="#F5F7FA", 
                text_color="#6B7C93", 
                hover_color="#E0E4EA", 
                corner_radius=8, 
                anchor="w", 
                height=44, 
                command=lambda idx=i: self.switch_page(idx)
            )
            btn.pack(fill="x", padx=16, pady=2)
            self.nav_btns.append(btn)
        # Main content area
        self.content_frame = ctk.CTkFrame(self, fg_color="#F5F7FA")
        self.content_frame.pack(side="left", fill="both", expand=True)
        self.pages = [
            HSCodePage(self.content_frame),
            BuyerSearchPage(self.content_frame),
            DeepSeekBuyerResultsPage(self.content_frame),
            ApolloPage(self.content_frame),
            ApolloBuyerListPage(self.content_frame),
            ExportPage(self.content_frame),
            ctk.CTkFrame(self.content_frame, fg_color="#F5F7FA"),  # Placeholder for Settings
        ]
        for page in self.pages:
            page.place(relx=0, rely=0, relwidth=1, relheight=1)
            page.lower()
        self.switch_page(0)
    

    
    def maximize_window(self):
        """Maximize the window reliably"""
        self.state('zoomed')
        self.update()
    
    def switch_page(self, idx):
        for i, btn in enumerate(self.nav_btns):
            if i == idx:
                btn.configure(fg_color="#FFFFFF", text_color="#0078D4")
            else:
                btn.configure(fg_color="#F5F7FA", text_color="#6B7C93")
        for i, page in enumerate(self.pages):
            if i == idx:
                page.lift()
            else:
                page.lower()
    
    def on_closing(self):
        """Handle app shutdown"""
        task_manager.shutdown()
        self.quit()

class ApolloBuyerListPage(ctk.CTkFrame):
    def __init__(self, master):
        super().__init__(master, fg_color="#F5F7FA")
        self.selected_company = None
        self.current_page = 0
        self.items_per_page = 50
        self._build_ui()
        self.populate_table()

    def _build_ui(self):
        # Title
        title_frame = ctk.CTkFrame(self, fg_color="#F5F7FA")
        title_frame.pack(fill="x", pady=(24, 16), padx=32)
        ctk.CTkLabel(title_frame, text="Buyer List (Apollo)", font=("Poppins", 24, "bold"), text_color="#2E3A59").pack(side="left")
        # Search/filter row
        filter_frame = ctk.CTkFrame(self, fg_color="#F5F7FA")
        filter_frame.pack(fill="x", pady=(18, 0), padx=32)
        ctk.CTkLabel(filter_frame, text="Search:", font=("Poppins", 15), text_color="#6B7C93").pack(side="left", padx=(0, 8))
        self.search_var = tk.StringVar()
        self.search_var.trace("w", self.on_search_change)
        search_entry = ctk.CTkEntry(filter_frame, textvariable=self.search_var, placeholder_text="Search name, company, email...", width=300, font=("Poppins", 15))
        search_entry.pack(side="left", padx=(0, 12))
        # Company filter
        ctk.CTkLabel(filter_frame, text="Company:", font=("Poppins", 15), text_color="#6B7C93").pack(side="left", padx=(16, 8))
        self.company_filter_var = tk.StringVar(value="All")
        company_frame = ctk.CTkFrame(filter_frame, fg_color="transparent")
        company_frame.pack(side="left", padx=(0, 12))
        self.company_display = ctk.CTkEntry(company_frame, textvariable=self.company_filter_var, state="readonly", width=180, font=("Poppins", 15))
        self.company_display.pack(side="left")
        select_company_btn = ctk.CTkButton(company_frame, text="Select", fg_color="#4CAF50", hover_color="#388E3C", text_color="#FFFFFF", font=("Poppins", 12, "bold"), width=60, height=32, command=self.open_company_selector)
        select_company_btn.pack(side="left", padx=(8, 0))
        refresh_btn = ctk.CTkButton(filter_frame, text="Refresh", fg_color="#4CAF50", hover_color="#388E3C", text_color="#FFFFFF", font=("Poppins", 15), corner_radius=8, width=100, command=self.populate_table)
        refresh_btn.pack(side="left", padx=(0, 8))
        clear_all_btn = ctk.CTkButton(filter_frame, text="Clear All", fg_color="#B0BEC5", hover_color="#90A4AE", text_color="#FFFFFF", font=("Poppins", 15), corner_radius=8, width=120, command=self.clear_all_filters)
        clear_all_btn.pack(side="left")
        # Table frame
        table_frame = ctk.CTkFrame(self, fg_color="#FFFFFF", corner_radius=16)
        table_frame.pack(fill="both", expand=True, padx=32, pady=(18, 0))
        style = ttk.Style()
        style.configure("ApolloBuyer.Treeview", font=("Poppins", 13), rowheight=32, background="#FFFFFF", fieldbackground="#FFFFFF", foreground="#2E3A59")
        style.configure("ApolloBuyer.Treeview.Heading", font=("Poppins", 14, "bold"), background="#F5F7FA", foreground="#0078D4")
        style.map("ApolloBuyer.Treeview", background=[("selected", "#1E3A8A")], foreground=[("selected", "#FFFFFF")])
        self.table = ttk.Treeview(table_frame, columns=("id", "name", "title", "email", "linkedin", "company_name", "source", "created_at"), show="headings", style="ApolloBuyer.Treeview")
        self.table.heading("id", text="ID")
        self.table.heading("name", text="Name")
        self.table.heading("title", text="Title")
        self.table.heading("email", text="Email")
        self.table.heading("linkedin", text="LinkedIn")
        self.table.heading("company_name", text="Company Name")
        self.table.heading("source", text="Source")
        self.table.heading("created_at", text="Created At")
        self.table.column("id", width=60, anchor="center")
        self.table.column("name", width=150, anchor="w")
        self.table.column("title", width=200, anchor="w")
        self.table.column("email", width=200, anchor="w")
        self.table.column("linkedin", width=150, anchor="w")
        self.table.column("company_name", width=200, anchor="w")
        self.table.column("source", width=100, anchor="center")
        self.table.column("created_at", width=140, anchor="center")
        self.table.pack(fill="both", expand=True, padx=8, pady=8)
        # Pagination controls
        self.pagination_frame = ctk.CTkFrame(self, fg_color="#F5F7FA")
        self.pagination_frame.pack(fill="x", padx=32, pady=(0, 8))
        self.prev_btn = ctk.CTkButton(self.pagination_frame, text="Previous", command=self.prev_page, width=80)
        self.next_btn = ctk.CTkButton(self.pagination_frame, text="Next", command=self.next_page, width=80)
        self.page_label = ctk.CTkLabel(self.pagination_frame, text="")
        self.prev_btn.pack(side="left")
        self.page_label.pack(side="left", padx=8)
        self.next_btn.pack(side="left")
        # Action buttons below table
        action_frame = ctk.CTkFrame(self, fg_color="#F5F7FA")
        action_frame.pack(fill="x", padx=32, pady=(8, 24))
        export_btn = ctk.CTkButton(action_frame, text="Export Results", fg_color="#0078D4", hover_color="#005A9E", text_color="#FFFFFF", font=("Poppins", 15, "bold"), corner_radius=8, width=150, command=self.export_results)
        export_btn.pack(side="right", padx=(0, 12))
        edit_btn = ctk.CTkButton(action_frame, text="Edit Selected", fg_color="#4CAF50", hover_color="#388E3C", text_color="#FFFFFF", font=("Poppins", 15), corner_radius=8, width=140, command=self.edit_selected)
        edit_btn.pack(side="right", padx=(0, 12))
        delete_btn = ctk.CTkButton(action_frame, text="Delete Selected", fg_color="#F44336", hover_color="#D32F2F", text_color="#FFFFFF", font=("Poppins", 15), corner_radius=8, width=140, command=self.delete_selected)
        delete_btn.pack(side="right", padx=(0, 12))

    def open_company_selector(self):
        import GUI_db
        companies = sorted(set(c['company_name'] for c in GUI_db.get_all_contacts() if c.get('company_name')))
        dialog = CompanySelectorDialog(self, companies)
        self.wait_window(dialog)
        selected_company = dialog.get_selected()
        if selected_company:
            self.company_filter_var.set(selected_company)
            self.current_page = 0
            self.populate_table()

    def clear_all_filters(self):
        """Clear both search term and company filter"""
        self.search_var.set("")
        self.company_filter_var.set("All")
        self.current_page = 0
        self.populate_table()

    def populate_table(self):
        """Populate the table with Apollo contacts, filtered and paginated"""
        import GUI_db
        search_term = self.search_var.get().strip().lower()
        company_filter = self.company_filter_var.get()
        results = GUI_db.get_all_contacts()
        # Filter by company
        if company_filter and company_filter != "All":
            results = [c for c in results if c.get('company_name', '') == company_filter]
        # Filter by search
        if search_term:
            filtered = []
            for entry in results:
                if (search_term in str(entry.get('name', '')).lower() or
                    search_term in str(entry.get('company_name', '')).lower() or
                    search_term in str(entry.get('email', '')).lower()):
                    filtered.append(entry)
            results = filtered
        # Sort by id ascending
        results = sorted(results, key=lambda x: x.get('id', 0))
        # Pagination
        total_pages = max(1, (len(results) - 1) // self.items_per_page + 1)
        start = self.current_page * self.items_per_page
        end = start + self.items_per_page
        page_results = results[start:end]
        for row in self.table.get_children():
            self.table.delete(row)
        for entry in page_results:
            self.table.insert("", "end", iid=entry['id'], values=(
                entry.get('id', ''),
                entry.get('name', ''),
                entry.get('title', ''),
                entry.get('email', ''),
                entry.get('linkedin', ''),
                entry.get('company_name', ''),
                entry.get('source', ''),
                entry.get('created_at', '')
            ))
        self.page_label.configure(text=f"Page {self.current_page + 1} of {total_pages}")
        self.prev_btn.configure(state="normal" if self.current_page > 0 else "disabled")
        self.next_btn.configure(state="normal" if (self.current_page + 1) * self.items_per_page < len(results) else "disabled")

    def next_page(self):
        self.current_page += 1
        self.populate_table()

    def prev_page(self):
        if self.current_page > 0:
            self.current_page -= 1
            self.populate_table()

    def on_search_change(self, *args):
        self.current_page = 0
        self.populate_table()

    def clear_search(self):
        self.search_var.set("")
        self.current_page = 0
        self.populate_table()

    def edit_selected(self):
        selected = self.table.selection()
        if not selected:
            messagebox.showwarning("Edit Contact", "Please select a row to edit.")
            return
        contact_id = int(selected[0])
        import GUI_db
        record = next((c for c in GUI_db.get_all_contacts() if c['id'] == contact_id), None)
        if not record:
            messagebox.showerror("Not Found", "Selected contact not found.")
            return
        dialog = ctk.CTkToplevel(self)
        dialog.title("Edit Contact")
        dialog.geometry("600x600")
        dialog.grab_set()
        dialog.transient(self.winfo_toplevel())
        ctk.CTkLabel(dialog, text="Edit Contact", font=("Poppins", 20, "bold"), text_color="#2E3A59").pack(pady=(24, 16))
        form_frame = ctk.CTkFrame(dialog, fg_color="#F5F7FA")
        form_frame.pack(fill="both", expand=True, padx=24, pady=(0, 16))
        # Name
        ctk.CTkLabel(form_frame, text="Name:", font=("Poppins", 15)).pack(anchor="w", padx=16, pady=(16, 4))
        name_var = tk.StringVar(value=record.get('name', ''))
        name_entry = ctk.CTkEntry(form_frame, textvariable=name_var, font=("Poppins", 14))
        name_entry.pack(fill="x", padx=16, pady=(0, 12))
        # Title
        ctk.CTkLabel(form_frame, text="Title:", font=("Poppins", 15)).pack(anchor="w", padx=16, pady=(0, 4))
        title_var = tk.StringVar(value=record.get('title', ''))
        title_entry = ctk.CTkEntry(form_frame, textvariable=title_var, font=("Poppins", 14))
        title_entry.pack(fill="x", padx=16, pady=(0, 12))
        # Email
        ctk.CTkLabel(form_frame, text="Email:", font=("Poppins", 15)).pack(anchor="w", padx=16, pady=(0, 4))
        email_var = tk.StringVar(value=record.get('email', ''))
        email_entry = ctk.CTkEntry(form_frame, textvariable=email_var, font=("Poppins", 14))
        email_entry.pack(fill="x", padx=16, pady=(0, 12))
        # LinkedIn
        ctk.CTkLabel(form_frame, text="LinkedIn:", font=("Poppins", 15)).pack(anchor="w", padx=16, pady=(0, 4))
        linkedin_var = tk.StringVar(value=record.get('linkedin', ''))
        linkedin_entry = ctk.CTkEntry(form_frame, textvariable=linkedin_var, font=("Poppins", 14))
        linkedin_entry.pack(fill="x", padx=16, pady=(0, 12))
        # Company Name
        ctk.CTkLabel(form_frame, text="Company Name:", font=("Poppins", 15)).pack(anchor="w", padx=16, pady=(0, 4))
        company_name_var = tk.StringVar(value=record.get('company_name', ''))
        company_name_entry = ctk.CTkEntry(form_frame, textvariable=company_name_var, font=("Poppins", 14))
        company_name_entry.pack(fill="x", padx=16, pady=(0, 12))
        # Buttons
        button_frame = ctk.CTkFrame(dialog, fg_color="#F5F7FA")
        button_frame.pack(fill="x", padx=24, pady=(0, 24))
        def on_save():
            updated_fields = {
                'name': name_var.get().strip(),
                'title': title_var.get().strip(),
                'email': email_var.get().strip(),
                'linkedin': linkedin_var.get().strip(),
                'company_name': company_name_var.get().strip(),
            }
            if not updated_fields['name'] or not updated_fields['company_name']:
                messagebox.showwarning("Missing Data", "Please fill in Name and Company Name.")
                return
            if GUI_db.update_contact(contact_id, updated_fields):
                dialog.destroy()
                self.populate_table()
                messagebox.showinfo("Success", "Contact updated successfully.")
            else:
                messagebox.showerror("Error", "Failed to update contact.")
        ctk.CTkButton(button_frame, text="Save", fg_color="#0078D4", text_color="#FFFFFF", font=("Poppins", 15, "bold"), command=on_save).pack(side="left", padx=(0, 8))
        ctk.CTkButton(button_frame, text="Cancel", fg_color="#B0BEC5", text_color="#FFFFFF", font=("Poppins", 15), command=dialog.destroy).pack(side="right")

    def delete_selected(self):
        selected = self.table.selection()
        if not selected:
            messagebox.showwarning("Delete Contact", "Please select a row to delete.")
            return
        contact_id = int(selected[0])
        if messagebox.askyesno("Delete Contact", "Are you sure you want to delete this contact?"):
            if GUI_db.delete_contact(contact_id):
                self.populate_table()
                messagebox.showinfo("Success", "Contact deleted successfully.")
            else:
                messagebox.showerror("Error", "Failed to delete contact.")

    def export_results(self):
        import csv
        from tkinter import filedialog
        results = []
        for item in self.table.get_children():
            values = self.table.item(item)['values']
            results.append(values)
        if not results:
            messagebox.showwarning("No Data", "No results to export.")
            return
        filename = filedialog.asksaveasfilename(
            defaultextension=".csv",
            filetypes=[("CSV files", "*.csv"), ("All files", "*.*")],
            title="Export Buyer List (Apollo)",
            initialfile="apollo_buyer_list.csv"
        )
        if filename:
            with open(filename, 'w', newline='', encoding='utf-8') as csvfile:
                writer = csv.writer(csvfile)
                headers = ["ID", "Name", "Title", "Email", "LinkedIn", "Company Name", "Source", "Created At"]
                writer.writerow(headers)
                for row in results:
                    writer.writerow(row)
            messagebox.showinfo("Export Complete", f"Results exported to {filename}")

if __name__ == "__main__":
    app = MainApp()
    app.protocol("WM_DELETE_WINDOW", app.on_closing)
    app.mainloop() 